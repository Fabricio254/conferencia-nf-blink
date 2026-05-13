"""
Conferência NF de Entrada — ERP vs Receita Federal
Streamlit App — roda localmente na rede da empresa
  Iniciar: venv\Scripts\streamlit.exe run app.py
"""

import os
import json
import io
import zipfile
import configparser
import subprocess
import fdb
import xml.etree.ElementTree as ET
from decimal import Decimal, ROUND_HALF_UP
from datetime import date
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import pandas as pd
import streamlit as st
_VERSAO_MARCA = "v14-2026-05-04-17:55"

try:
    from sefaz_distdfe import (
        consultar_distdfe, filtrar_nfes, filtrar_nfes_saida, info_certificado,
        resetar_nsu, salvar_ultimo_nsu, UF_CODIGOS,
        carregar_cache, get_info_cache,
        _ConsumoIndevido,
        extrair_res_nfe_ch, manifestar_ciencia,
        manifestar_evento, carregar_manifestacoes,
        baixar_proc_nfe_por_chave,
    )
    SEFAZ_OK = True
except ImportError:
    SEFAZ_OK = False

# ─── PATHS ────────────────────────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
DB_CONFIG    = os.path.join(BASE_DIR, 'db_config.json')
CERT_CONFIG  = os.path.join(BASE_DIR, 'cert_config.json')
LOJAS_CONFIG = os.path.join(BASE_DIR, 'lojas_config.json')
EMAIL_CONFIG = os.path.join(BASE_DIR, 'email_config.json')

FB_DLLS = [
    r'C:\Program Files (x86)\VisualControl\bin\fbclient25.dll',
    r'C:\Program Files\Firebird\Firebird_2_5\fbclient.dll',
    r'C:\Program Files (x86)\Firebird\Firebird_2_5\fbclient.dll',
    r'C:\Program Files\Firebird\Firebird_5_0\fbclient.dll',
    r'C:\Program Files\Firebird\Firebird_4_0\fbclient.dll',
    r'C:\Program Files\Firebird\Firebird_3_0\fbclient.dll',
]

NS_NFE = '{http://www.portalfiscal.inf.br/nfe}'

STATUS_OK          = 'OK'
STATUS_DIVERGENCIA = 'DIVERGÊNCIA'
STATUS_SO_ERP      = 'SOMENTE NO ERP'
STATUS_SO_XML      = 'SOMENTE NO XML'
STATUS_RESUMO      = 'RESUMO SEFAZ'

# ─── CONFIG ───────────────────────────────────────────────────────────────────

def _load_json(path, defaults):
    if os.path.exists(path):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return {**defaults, **json.load(f)}
        except Exception:
            pass
    return dict(defaults)

def _save_json(path, data):
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def load_db_config():
    return _load_json(DB_CONFIG, {
        'host'    : '192.168.15.240',
        'port'    : 3051,
        'database': r'C:\Program Files (x86)\VisualControl\Database\ERP_VCONTROL_BLINK.FDB',
        'user'    : 'SYSDBA',
        'password': 'masterkey',
    })

def load_email_config():
    return _load_json(EMAIL_CONFIG, {
        'remetente'   : 'blinkcpd@gmail.com',
        'senha_app'   : '',
        'destinatario': '',
        'brevo_api_key': '',
    })


# Caminhos padrão onde o VisualControl guarda suas configurações
_VC_INI_PATHS = [
    r'C:\Program Files (x86)\VisualControl\VisualControl.ini',
    r'C:\Program Files\VisualControl\VisualControl.ini',
    r'C:\VisualControl\VisualControl.ini',
]

def _decodificar_senha_vc(senha_raw):
    """
    O VisualControl armazena a senha como string hex (ex: 5D4343364D5E5F2941).
    Decodifica para o texto real. Se não for hex válido, retorna o valor original.
    """
    t = senha_raw.strip()
    if len(t) % 2 == 0 and all(c in '0123456789ABCDEFabcdef' for c in t):
        try:
            return bytes.fromhex(t).decode('latin-1')
        except Exception:
            pass
    return t

def ler_config_visualcontrol():
    """
    Lê o VisualControl.ini e retorna dict com os campos de conexão Firebird,
    ou None se o arquivo não for encontrado.
    """
    for caminho in _VC_INI_PATHS:
        if os.path.exists(caminho):
            cfg = configparser.ConfigParser()
            # O VisualControl.ini não tem espaços ao redor do '='
            cfg.read(caminho, encoding='latin-1')
            secao = None
            for s in ('BancoDados', 'bancoDados', 'BANCODADOS'):
                if cfg.has_section(s):
                    secao = s
                    break
            if secao is None:
                return None
            servidor = cfg.get(secao, 'Servidor', fallback='localhost').strip()
            bd       = cfg.get(secao, 'BD',       fallback='').strip()
            usuario  = cfg.get(secao, 'Usuario',  fallback='SYSDBA').strip()
            senha    = _decodificar_senha_vc(cfg.get(secao, 'Senha', fallback='').strip())
            # "localhost" no ini significa o próprio servidor — mas para o Streamlit
            # rodando em outra máquina precisamos do IP real.
            # Deixamos como veio e o usuário pode ajustar.
            return {
                'host'    : servidor,
                'port'    : 3051,
                'database': bd,
                'user'    : usuario,
                'password': senha,
                '_ini_path': caminho,
            }
    return None

def load_cert_config():
    return _load_json(CERT_CONFIG, {
        'pfx_path'    : '',
        'senha'       : '',
        'cnpj'        : '',
        'uf'          : 32,
        'pasta_nfce'  : '',
    })

def load_lojas_config():
    """Carrega lojas_config.json com todas as lojas cadastradas."""
    if not os.path.exists(LOJAS_CONFIG):
        return []
    try:
        with open(LOJAS_CONFIG, 'r', encoding='utf-8') as _f:
            return json.load(_f)
    except Exception:
        return []

def _listar_certs_windows():
    """Lista certificados A1 com chave privada do store CurrentUser\\My.
    Retorna lista de dicts ordenada pelo mais recentemente instalado."""
    try:
        ps_script = (
            'Get-ChildItem Cert:\\CurrentUser\\My '
            '| Where-Object { $_.HasPrivateKey -eq $true } '
            '| Select-Object Subject,Thumbprint,'
            '@{n="NotAfter";e={$_.NotAfter.ToString("yyyy-MM-dd")}},'
            '@{n="NotBefore";e={$_.NotBefore.ToString("yyyy-MM-dd")}} '
            '| ConvertTo-Json -Compress'
        )
        res = subprocess.run(
            ['powershell', '-NoProfile', '-NonInteractive', '-Command', ps_script],
            capture_output=True, text=True, timeout=10
        )
        if res.returncode != 0 or not res.stdout.strip():
            return []
        data = json.loads(res.stdout.strip())
        if isinstance(data, dict):
            data = [data]
        # Prefere certs com CNPJ no Subject (e-CNPJ NF-e)
        nfe = [c for c in data if 'CNPJ' in (c.get('Subject') or '')]
        outros = [c for c in data if c not in nfe]
        lista = nfe + outros
        # Ordena pelo NotBefore (instalação) mais recente
        lista.sort(key=lambda c: c.get('NotBefore', ''), reverse=True)
        return lista
    except Exception:
        return []

def _exportar_cert_windows(thumbprint, senha_export, dest_path):
    """Exporta certificado do Windows store para .pfx. Retorna (True, '') ou (False, erro)."""
    try:
        # Escapar aspas na senha para evitar injeção no comando
        senha_safe = senha_export.replace('"', '`"')
        dest_safe  = dest_path.replace('"', '`"')
        ps_script = (
            f'$cert = Get-Item "Cert:\\CurrentUser\\My\\{thumbprint}"; '
            f'$pwd = ConvertTo-SecureString -String "{senha_safe}" -Force -AsPlainText; '
            f'Export-PfxCertificate -Cert $cert -FilePath "{dest_safe}" -Password $pwd | Out-Null; '
            f'Write-Host "EXPORTADO_OK"'
        )
        res = subprocess.run(
            ['powershell', '-NoProfile', '-NonInteractive', '-Command', ps_script],
            capture_output=True, text=True, timeout=20
        )
        if 'EXPORTADO_OK' in res.stdout:
            return True, ''
        return False, (res.stderr or res.stdout).strip()
    except Exception as e:
        return False, str(e)

def _buscar_dados_loja(db_cfg):
    """Conecta ao ERP e retorna (cnpj_limpo, uf_sigla) da primeira loja, ou (None, None)."""
    try:
        con = conectar_erp(db_cfg)
        cur = con.cursor()
        cnpj = None
        uf_sigla = None
        try:
            cur.execute('SELECT FIRST 1 CPFCGC FROM LOJA ORDER BY ID')
            row = cur.fetchone()
            if row and row[0]:
                cnpj = _limpar_cnpj(str(row[0]))
        except Exception:
            pass
        for col in ('UF', 'CODUF', 'COD_UF', 'ESTADO'):
            try:
                cur.execute(f'SELECT FIRST 1 {col} FROM LOJA ORDER BY ID')
                row = cur.fetchone()
                if row and row[0]:
                    uf_sigla = str(row[0]).strip().upper()
                    break
            except Exception:
                continue
        con.close()
        return cnpj, uf_sigla
    except Exception:
        return None, None

# ─── FIREBIRD ─────────────────────────────────────────────────────────────────

def conectar_erp(cfg):
    for dll in FB_DLLS:
        if os.path.exists(dll):
            fdb.load_api(dll)
            break
    else:
        raise FileNotFoundError('fbclient.dll não encontrada. Instale o Firebird Client.')
    return fdb.connect(
        host=cfg['host'], port=int(cfg['port']),
        database=cfg['database'],
        user=cfg['user'], password=cfg['password'],
        charset='UTF8',
    )

def conectar_erp_loja(loja, db_cfg_global):
    """
    Conecta no banco local da loja se ela tiver 'db_host' configurado.
    Caso contrário usa o banco central (db_cfg_global).
    """
    if loja and loja.get('db_host') and loja.get('db_database'):
        cfg_local = {
            'host'    : loja['db_host'],
            'port'    : loja.get('db_port', 3050),
            'database': loja['db_database'],
            'user'    : loja.get('db_user', 'SYSDBA'),
            'password': loja.get('db_password', ''),
        }
        return conectar_erp(cfg_local)
    return conectar_erp(db_cfg_global)

def buscar_nfs_erp(con, data_ini, data_fim, usar_emissao=False, loja_id=None):
    cur = con.cursor()
    campo_data = 'e.EMISSAO' if usar_emissao else 'e.DATA_ENTRADA'
    # Filtro por loja via prefixo do ID (ex: loja 17 → IDs 170000000..179999999)
    filtro_loja = ''
    if loja_id is not None:
        prefixo = int(loja_id)
        if prefixo < 10:
            # Loja 1..9: prefixo ex "10000000" (8 dígitos) a "19999999"
            id_min = prefixo * 10000000
            id_max = (prefixo + 1) * 10000000 - 1
        else:
            # Loja 10..99: prefixo ex "100000000" (9 dígitos) a "109999999"
            id_min = prefixo * 10000000
            id_max = (prefixo + 1) * 10000000 - 1
        filtro_loja = f'AND e.ID BETWEEN {id_min} AND {id_max}'
    cur.execute(f"""
        SELECT
            e.NOTAFISCAL                              AS NUMERO,
            TRIM(e.SERIE)                             AS SERIE,
            e.DATA_ENTRADA                            AS DT_ENTRADA,
            e.EMISSAO                                 AS DT_EMISSAO,
            ''                                        AS CHAVE,
            TRIM(cl.NOME)                             AS FORNECEDOR,
            TRIM(cl.CPFCGC)                           AS CNPJ_FORN,
            CAST(e.TOTAL_NF    AS DOUBLE PRECISION)   AS TOTAL_NF,
            CAST(e.TOTAL_PROD  AS DOUBLE PRECISION)   AS TOTAL_PROD,
            CAST(COALESCE(e.TOTAL_IPI,   0) AS DOUBLE PRECISION) AS TOTAL_IPI,
            CAST(COALESCE(e.TOTAL_FRETE, 0) AS DOUBLE PRECISION) AS TOTAL_FRETE,
            CAST(COALESCE(e.VALOR_ICMS,  0) AS DOUBLE PRECISION) AS TOTAL_ICMS,
            CAST(COALESCE(e.VALOR_ICMS_SUBS, 0) AS DOUBLE PRECISION) AS TOTAL_ST,
            CAST(COALESCE(e.TOTAL_NF, 0) - COALESCE(e.TOTAL_LIQUIDO, e.TOTAL_NF, 0)
                 AS DOUBLE PRECISION) AS DESCONTO,
            ''                                        AS LOJA
        FROM ENTRADA_NF e
        INNER JOIN CLIFOR cl ON cl.ID = e.ID_FORNECEDOR
        WHERE COALESCE(e.STATUS, 'N') <> 'C'
          AND {campo_data} BETWEEN ? AND ?
          {filtro_loja}
        ORDER BY e.DATA_ENTRADA DESC, e.NOTAFISCAL
    """, (data_ini, data_fim))
    cols = [d[0] for d in cur.description]
    rows = []
    for row in cur.fetchall():
        r = dict(zip(cols, row))
        r['CHAVE']      = (r['CHAVE'] or '').strip()
        r['CNPJ_FORN']  = _limpar_cnpj(r['CNPJ_FORN'])
        r['TOTAL_NF']   = _moeda(r['TOTAL_NF'])
        r['TOTAL_PROD'] = _moeda(r['TOTAL_PROD'])
        r['TOTAL_IPI']  = _moeda(r['TOTAL_IPI'])
        r['TOTAL_FRETE']= _moeda(r['TOTAL_FRETE'])
        r['TOTAL_ICMS'] = _moeda(r['TOTAL_ICMS'])
        r['TOTAL_ST']   = _moeda(r['TOTAL_ST'])
        r['DESCONTO']   = _moeda(r['DESCONTO'])
        rows.append(r)
    cur.close()
    return rows

def buscar_nfs_saida_erp(con, data_ini, data_fim, cnpj_loja=None, loja_id=None):
    """Busca NF-e de saída (modelo 55) no ERP, incluindo o XML do banco."""
    cur = con.cursor()
    params = [data_ini, data_fim]
    if loja_id is not None:
        filtro_loja = 'AND nf.ID_LOJA = ?'
        params.append(int(loja_id))
    elif cnpj_loja:
        cnpj_limpo = _limpar_cnpj(cnpj_loja)
        filtro_loja = f"AND nf.ID_LOJA = (SELECT FIRST 1 ID FROM LOJA WHERE REPLACE(REPLACE(REPLACE(CPFCGC,'.',''),'-',''),'/','') = '{cnpj_limpo}')"
    else:
        filtro_loja = ''
    cur.execute(f"""
        SELECT
            nf.ID_COD            AS NUMERO,
            TRIM(nf.SERIE)       AS SERIE,
            nf.DATA              AS DT_SAIDA,
            nf.EMISSAO           AS DT_EMISSAO,
            nf.CHAVE_ACESSO_NFE  AS CHAVE,
            TRIM(cl.NOME)        AS CLIENTE,
            TRIM(cl.CPFCGC)      AS CNPJ_CLI,
            CAST(nf.TOTAL_NF      AS DOUBLE PRECISION) AS TOTAL_NF,
            CAST(nf.TOTAL_PROD    AS DOUBLE PRECISION) AS TOTAL_PROD,
            CAST(nf.TOTAL_IPI     AS DOUBLE PRECISION) AS TOTAL_IPI,
            CAST(nf.TOTAL_FRETE   AS DOUBLE PRECISION) AS TOTAL_FRETE,
            CAST(nf.TOTAL_ICMS    AS DOUBLE PRECISION) AS TOTAL_ICMS,
            CAST(nf.TOTAL_ICMS_ST AS DOUBLE PRECISION) AS TOTAL_ST,
            CAST(nf.DESCONTO_VALOR AS DOUBLE PRECISION) AS DESCONTO,
            TRIM(lj.NOME) AS LOJA,
            nf.XML               AS XML_BLOB
        FROM NFISCAL nf
        INNER JOIN CLIFOR cl ON cl.ID = nf.ID_CLIFOR
        INNER JOIN LOJA   lj ON lj.ID = nf.ID_LOJA
        WHERE nf.TIPO_ES  = 'S'
          AND COALESCE(nf.STATUS, 'N') <> 'C'
          AND nf.DATA BETWEEN ? AND ?
          {filtro_loja}
        ORDER BY nf.DATA DESC, nf.ID_COD
    """, params)
    cols = [d[0] for d in cur.description]
    rows = []
    for row in cur.fetchall():
        r = dict(zip(cols, row))
        r['CHAVE']      = (r['CHAVE'] or '').strip()
        r['CNPJ_CLI']   = _limpar_cnpj(r['CNPJ_CLI'])
        r['TOTAL_NF']   = _moeda(r['TOTAL_NF'])
        r['TOTAL_PROD'] = _moeda(r['TOTAL_PROD'])
        r['TOTAL_IPI']  = _moeda(r['TOTAL_IPI'])
        r['TOTAL_FRETE']= _moeda(r['TOTAL_FRETE'])
        r['TOTAL_ICMS'] = _moeda(r['TOTAL_ICMS'])
        r['TOTAL_ST']   = _moeda(r['TOTAL_ST'])
        r['DESCONTO']   = _moeda(r['DESCONTO'])
        # Ler BLOB XML
        blob = r.pop('XML_BLOB', None)
        xml_str = ''
        if blob is not None:
            try:
                raw = blob.read() if hasattr(blob, 'read') else blob
                xml_str = raw.decode('utf-8', errors='ignore') if isinstance(raw, bytes) else (raw or '')
                xml_str = xml_str.strip()
            except Exception:
                xml_str = ''
        r['XML'] = xml_str
        rows.append(r)
    cur.close()
    return rows


def buscar_nfs_nfiscal_entrada(con, data_ini, data_fim, loja_id, usar_emissao=False):
    """Busca NF-e de entrada via tabela NFISCAL (lojas 02, 10, 14 que lançam entrada no NFISCAL)."""
    cur = con.cursor()
    campo_data = 'nf.EMISSAO' if usar_emissao else 'nf.DATA'
    cur.execute(f"""
        SELECT
            nf.ID_COD                                   AS NUMERO,
            TRIM(nf.SERIE)                              AS SERIE,
            nf.DATA                                     AS DT_ENTRADA,
            nf.EMISSAO                                  AS DT_EMISSAO,
            TRIM(COALESCE(nf.CHAVE_ACESSO_NFE, ''))     AS CHAVE,
            TRIM(cl.NOME)                               AS FORNECEDOR,
            TRIM(cl.CPFCGC)                             AS CNPJ_FORN,
            CAST(nf.TOTAL_NF       AS DOUBLE PRECISION) AS TOTAL_NF,
            CAST(nf.TOTAL_PROD     AS DOUBLE PRECISION) AS TOTAL_PROD,
            CAST(COALESCE(nf.TOTAL_IPI,    0) AS DOUBLE PRECISION) AS TOTAL_IPI,
            CAST(COALESCE(nf.TOTAL_FRETE,  0) AS DOUBLE PRECISION) AS TOTAL_FRETE,
            CAST(COALESCE(nf.TOTAL_ICMS,   0) AS DOUBLE PRECISION) AS TOTAL_ICMS,
            CAST(COALESCE(nf.TOTAL_ICMS_ST,0) AS DOUBLE PRECISION) AS TOTAL_ST,
            CAST(COALESCE(nf.DESCONTO_VALOR,0) AS DOUBLE PRECISION) AS DESCONTO,
            TRIM(lj.NOME)                               AS LOJA,
            nf.XML                                      AS XML_BLOB
        FROM NFISCAL nf
        INNER JOIN CLIFOR cl ON cl.ID = nf.ID_CLIFOR
        INNER JOIN LOJA   lj ON lj.ID = nf.ID_LOJA
        WHERE nf.TIPO_ES = 'E'
          AND nf.ID_LOJA = {int(loja_id)}
          AND COALESCE(nf.STATUS, 'N') <> 'C'
          AND {campo_data} BETWEEN ? AND ?
        ORDER BY nf.DATA DESC, nf.ID_COD
    """, (data_ini, data_fim))
    cols = [d[0] for d in cur.description]
    rows = []
    for row in cur.fetchall():
        r = dict(zip(cols, row))
        r['CHAVE']      = (r['CHAVE'] or '').strip()
        r['CNPJ_FORN']  = _limpar_cnpj(r['CNPJ_FORN'])
        r['TOTAL_NF']   = _moeda(r['TOTAL_NF'])
        r['TOTAL_PROD'] = _moeda(r['TOTAL_PROD'])
        r['TOTAL_IPI']  = _moeda(r['TOTAL_IPI'])
        r['TOTAL_FRETE']= _moeda(r['TOTAL_FRETE'])
        r['TOTAL_ICMS'] = _moeda(r['TOTAL_ICMS'])
        r['TOTAL_ST']   = _moeda(r['TOTAL_ST'])
        r['DESCONTO']   = _moeda(r['DESCONTO'])
        # Ler BLOB XML
        blob = r.pop('XML_BLOB', None)
        xml_str = ''
        if blob is not None:
            try:
                raw = blob.read() if hasattr(blob, 'read') else blob
                xml_str = raw.decode('utf-8', errors='ignore') if isinstance(raw, bytes) else (raw or '')
                xml_str = xml_str.strip()
            except Exception:
                xml_str = ''
        r['XML'] = xml_str
        rows.append(r)
    cur.close()
    return rows


def conferir_saida(nfs_erp, nfs_xml):
    """Cruza NF-e de saída do ERP com XMLs da Receita (onde a empresa é emitente)."""
    xml_por_chave = {}
    xml_por_ident = {}
    xml_usados    = set()
    for xml in nfs_xml:
        if xml['CHAVE']:
            xml_por_chave[xml['CHAVE']] = xml
        ident = (str(xml['NUMERO']).strip(), str(xml['SERIE']).strip())
        xml_por_ident[ident] = xml

    resultados = []
    campos_val = ['TOTAL_NF', 'TOTAL_PROD', 'TOTAL_IPI',
                  'TOTAL_FRETE', 'TOTAL_ICMS', 'TOTAL_ST', 'DESCONTO']

    for erp in nfs_erp:
        chave_erp = (erp['CHAVE'] or '').replace('NFe', '')
        xml = xml_por_chave.get(chave_erp) if chave_erp else None
        if xml is None:
            ident = (str(erp['NUMERO'] or '').strip(), str(erp['SERIE'] or '').strip())
            xml = xml_por_ident.get(ident)
        if xml:
            xml_usados.add(xml['CHAVE'])
            divs = []
            for c in campos_val:
                v_erp = float(erp.get(c) or 0)
                v_xml = float(xml.get(c) or 0)
                if abs(v_erp - v_xml) > 0.05:
                    divs.append(f'{c}: ERP={v_erp:.2f} XML={v_xml:.2f}')
            status = STATUS_DIVERGENCIA if divs else STATUS_OK
            resultados.append({
                'Status': status, 'Número': erp['NUMERO'], 'Série': erp['SERIE'],
                'Dt. Entrada': str(erp['DT_SAIDA']) if erp['DT_SAIDA'] else '',
                'Dt. Emissão': str(erp['DT_EMISSAO']) if erp['DT_EMISSAO'] else '',
                'Fornecedor ERP': erp['CLIENTE'], 'Fornecedor XML': xml['DESTINATARIO'],
                'CNPJ': erp['CNPJ_CLI'], 'Loja': erp.get('LOJA', ''),
                'Total ERP': float(erp['TOTAL_NF']), 'Total XML': float(xml['TOTAL_NF']),
                'Diferença': float(erp['TOTAL_NF']) - float(xml['TOTAL_NF']),
                'Prod ERP': float(erp['TOTAL_PROD']), 'Prod XML': float(xml['TOTAL_PROD']),
                'IPI ERP': float(erp['TOTAL_IPI']), 'IPI XML': float(xml['TOTAL_IPI']),
                'Frete ERP': 0.0, 'Frete XML': 0.0,
                'ICMS ERP': float(erp['TOTAL_ICMS']), 'ICMS XML': float(xml['TOTAL_ICMS']),
                'ST ERP': float(erp['TOTAL_ST']), 'ST XML': float(xml['TOTAL_ST']),
                'Desc ERP': 0.0, 'Desc XML': 0.0,
                'Divergências': ' | '.join(divs),
                'Chave NF-e': erp['CHAVE'] or xml['CHAVE'],
                'Arquivo XML': xml.get('ARQUIVO', ''),
            })
        else:
            resultados.append({
                'Status': STATUS_SO_ERP, 'Número': erp['NUMERO'], 'Série': erp['SERIE'],
                'Dt. Entrada': str(erp['DT_SAIDA']) if erp['DT_SAIDA'] else '',
                'Dt. Emissão': str(erp['DT_EMISSAO']) if erp['DT_EMISSAO'] else '',
                'Fornecedor ERP': erp['CLIENTE'], 'Fornecedor XML': '',
                'CNPJ': erp['CNPJ_CLI'], 'Loja': erp.get('LOJA', ''),
                'Total ERP': float(erp['TOTAL_NF']), 'Total XML': 0.0,
                'Diferença': float(erp['TOTAL_NF']),
                'Prod ERP': float(erp['TOTAL_PROD']), 'Prod XML': 0.0,
                'IPI ERP': float(erp['TOTAL_IPI']), 'IPI XML': 0.0,
                'Frete ERP': 0.0, 'Frete XML': 0.0,
                'ICMS ERP': float(erp['TOTAL_ICMS']), 'ICMS XML': 0.0,
                'ST ERP': float(erp['TOTAL_ST']), 'ST XML': 0.0,
                'Desc ERP': 0.0, 'Desc XML': 0.0,
                'Divergências': 'NF não encontrada na Receita',
                'Chave NF-e': erp['CHAVE'],
                'Arquivo XML': '',
            })
    # XMLs na Receita mas não no ERP
    for xml in nfs_xml:
        if xml['CHAVE'] not in xml_usados:
            resultados.append({
                'Status': STATUS_SO_XML, 'Número': xml['NUMERO'], 'Série': xml['SERIE'],
                'Dt. Entrada': '', 'Dt. Emissão': xml['EMISSAO'],
                'Fornecedor ERP': '', 'Fornecedor XML': xml['DESTINATARIO'],
                'CNPJ': xml['CNPJ_DEST'], 'Loja': '',
                'Total ERP': 0.0, 'Total XML': float(xml['TOTAL_NF']),
                'Diferença': -float(xml['TOTAL_NF']),
                'Prod ERP': 0.0, 'Prod XML': float(xml['TOTAL_PROD']),
                'IPI ERP': 0.0, 'IPI XML': float(xml['TOTAL_IPI']),
                'Frete ERP': 0.0, 'Frete XML': 0.0,
                'ICMS ERP': 0.0, 'ICMS XML': float(xml['TOTAL_ICMS']),
                'ST ERP': 0.0, 'ST XML': float(xml['TOTAL_ST']),
                'Desc ERP': 0.0, 'Desc XML': 0.0,
                'Divergências': 'XML na Receita sem correspondência no ERP',
                'Chave NF-e': xml['CHAVE'],
                'Arquivo XML': xml.get('ARQUIVO', ''),
            })
    return resultados

# ─── UTILITÁRIOS ──────────────────────────────────────────────────────────────

def _moeda(v):
    if v is None:
        return Decimal('0.00')
    return Decimal(str(v)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

def _limpar_cnpj(cnpj):
    return ''.join(c for c in (cnpj or '') if c.isdigit())

def _txt(elem, tag, ns=NS_NFE):
    if elem is None:
        return ''
    node = elem.find(f'{ns}{tag}')
    return (node.text or '').strip() if node is not None else ''


def buscar_xml_no_cache(nfs_sem_xml, cnpj_loja):
    """
    Para cada NF sem XML do ERP, tenta encontrar o XML correspondente no cache
    local da SEFAZ, casando por (nNF + CNPJ emit + total vNF).
    Retorna dict: {idx_original: {'xml': str, 'chave': str, 'origem': str}}
    """
    if not SEFAZ_OK or not nfs_sem_xml:
        return {}
    import xml.etree.ElementTree as _ET
    _NS = 'http://www.portalfiscal.inf.br/nfe'

    docs = carregar_cache(_limpar_cnpj(cnpj_loja))
    if not docs:
        return {}

    # Monta índice: (nnf, cnpj_emit) → doc para lookup rápido
    indice = {}
    for doc in docs:
        xml_str = doc.get('xml', '')
        if not xml_str or 'procNFe' not in doc.get('schema', ''):
            continue
        try:
            root = _ET.fromstring(xml_str)
            inf = root.find(f'.//{{{_NS}}}infNFe')
            if inf is None:
                continue
            nnf  = (inf.findtext(f'{{{_NS}}}ide/{{{_NS}}}nNF') or '').strip()
            cnpj_e = _limpar_cnpj(inf.findtext(f'{{{_NS}}}emit/{{{_NS}}}CNPJ') or '')
            vnf  = inf.findtext(f'.//{{{_NS}}}vNF') or '0'
            chave = (inf.get('Id') or '').replace('NFe', '')
            indice.setdefault((nnf, cnpj_e), []).append({
                'xml': xml_str, 'chave': chave, 'vnf': Decimal(vnf)
            })
        except Exception:
            continue

    resultado = {}
    for i, nf in enumerate(nfs_sem_xml):
        nnf_erp   = str(nf.get('NUMERO', '')).strip().lstrip('0') or str(nf.get('NUMERO', ''))
        cnpj_emit = _limpar_cnpj(nf.get('CNPJ_EMIT') or nf.get('CNPJ_FORN') or '')
        total_erp = Decimal(str(nf.get('TOTAL_NF', 0)))
        candidatos = indice.get((nnf_erp, cnpj_emit), [])
        for c in candidatos:
            if abs(c['vnf'] - total_erp) < Decimal('0.02'):
                resultado[i] = {'xml': c['xml'], 'chave': c['chave'], 'origem': 'cache SEFAZ'}
                break
    return resultado


def enriquecer_entradas_com_cache(nfs_ei, cnpj_loja):
    """
    Para NFs de entrada onde o emitente É A PRÓPRIA LOJA (devoluções),
    o XML do ERP é correto — mantém.
    Para NFs onde o emitente é UM FORNECEDOR EXTERNO, descarta o XML do ERP
    e busca o XML correto no cache SEFAZ (destinatária = loja).
    Marca cada NF com 'XML_ORIGEM': 'ERP (devolução)' ou 'cache SEFAZ' ou 'não encontrado'.
    Retorna a lista modificada.
    """
    if not SEFAZ_OK or not nfs_ei:
        return nfs_ei
    import xml.etree.ElementTree as _ET
    _NS = 'http://www.portalfiscal.inf.br/nfe'
    cnpj_loja_limpo = _limpar_cnpj(cnpj_loja)

    # Separar compras de fornecedores externas (emitente ≠ própria loja)
    idx_compras = [i for i, nf in enumerate(nfs_ei)
                   if _limpar_cnpj(nf.get('CNPJ_FORN') or '') != cnpj_loja_limpo]

    if not idx_compras:
        for nf in nfs_ei:
            nf['XML_ORIGEM'] = 'ERP (devolução)'
        return nfs_ei

    # Marcar devoluções
    for i, nf in enumerate(nfs_ei):
        if i not in idx_compras:
            nf['XML_ORIGEM'] = 'ERP (devolução)'

    # Carregar cache SEFAZ e montar índice por chave e por (nNF, cnpj_emit)
    docs = carregar_cache(cnpj_loja_limpo)
    indice_chave = {}
    indice_nnf   = {}
    for doc in docs:
        xml_str = doc.get('xml', '')
        if not xml_str or 'procNFe' not in doc.get('schema', ''):
            continue
        try:
            root = _ET.fromstring(xml_str)
            inf  = root.find(f'.//{{{_NS}}}infNFe')
            if inf is None:
                continue
            chave  = (inf.get('Id') or '').replace('NFe', '')
            nnf    = (inf.findtext(f'{{{_NS}}}ide/{{{_NS}}}nNF') or '').strip()
            cnpj_e = _limpar_cnpj(inf.findtext(f'{{{_NS}}}emit/{{{_NS}}}CNPJ') or '')
            vnf    = Decimal(inf.findtext(f'.//{{{_NS}}}vNF') or '0')
            entry  = {'xml': xml_str, 'chave': chave, 'vnf': vnf}
            if chave:
                indice_chave[chave] = entry
            indice_nnf.setdefault((nnf, cnpj_e), []).append(entry)
        except Exception:
            continue

    for i in idx_compras:
        nf = nfs_ei[i]
        chave_erp  = _limpar_cnpj(nf.get('CHAVE') or '').replace('NFe', '')
        cnpj_forn  = _limpar_cnpj(nf.get('CNPJ_FORN') or '')
        nnf_erp    = str(nf.get('NUMERO', '')).strip().lstrip('0') or str(nf.get('NUMERO', ''))
        total_erp  = Decimal(str(nf.get('TOTAL_NF', 0)))

        found = None
        # 1) match por chave exata
        if chave_erp and chave_erp in indice_chave:
            found = indice_chave[chave_erp]
        # 2) match por nNF + cnpj_emit + valor
        if not found:
            for c in indice_nnf.get((nnf_erp, cnpj_forn), []):
                if abs(c['vnf'] - total_erp) < Decimal('0.02'):
                    found = c
                    break
        if found:
            nf['XML']        = found['xml']
            nf['CHAVE']      = found['chave']
            nf['XML_ORIGEM'] = 'cache SEFAZ'
        else:
            nf['XML']        = ''   # descarta XML de devolução que não é desta compra
            nf['XML_ORIGEM'] = 'não encontrado'

    return nfs_ei

def _fmt_brl(v):
    try:
        return f"R$ {float(v):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    except Exception:
        return str(v)

# ─── LEITURA DE XMLs ──────────────────────────────────────────────────────────

def _parse_nfe(root, nome):
    infNFe = root.find(f'.//{NS_NFE}infNFe')
    ns = NS_NFE
    if infNFe is None:
        infNFe = root.find('.//infNFe')
        ns = ''
    if infNFe is None:
        return None, 'Tag infNFe não encontrada'

    ide   = infNFe.find(f'{ns}ide')
    emit  = infNFe.find(f'{ns}emit')
    dest  = infNFe.find(f'{ns}dest')
    total = (infNFe.find(f'.//{ns}ICMSTot')
             or infNFe.find(f'{ns}total/{ns}ICMSTot')
             or infNFe.find(f'.//{NS_NFE}ICMSTot'))

    chave = infNFe.get('Id', '').replace('NFe', '')
    emissao = _txt(ide, 'dhEmi', ns) or _txt(ide, 'dEmi', ns)

    def val(tag):
        return _moeda(_txt(total, tag, ns) if total is not None else None)

    return {
        'ARQUIVO'      : nome,
        'CHAVE'        : chave,
        'NUMERO'       : _txt(ide, 'nNF', ns),
        'SERIE'        : _txt(ide, 'serie', ns),
        'EMISSAO'      : emissao[:10] if emissao else '',
        'CNPJ_EMIT'    : _limpar_cnpj(_txt(emit, 'CNPJ', ns)),
        'FORNECEDOR'   : _txt(emit, 'xNome', ns) or _txt(emit, 'xFant', ns),
        'CNPJ_DEST'    : _limpar_cnpj(_txt(dest, 'CNPJ', ns)) if dest is not None else '',
        'DESTINATARIO' : (_txt(dest, 'xNome', ns) or '') if dest is not None else '',
        'TOTAL_NF'     : val('vNF'),
        'TOTAL_PROD'   : val('vProd'),
        'TOTAL_IPI'    : val('vIPI'),
        'TOTAL_FRETE'  : val('vFrete'),
        'TOTAL_ICMS'   : val('vICMS'),
        'TOTAL_ST'     : val('vST'),
        'DESCONTO'     : val('vDesc'),
    }, None

def ler_xml_arquivo(caminho):
    try:
        return _parse_nfe(ET.parse(caminho).getroot(), os.path.basename(caminho))
    except ET.ParseError as e:
        return None, f'XML inválido: {e}'

def ler_xml_string(xml_str, nome):
    try:
        return _parse_nfe(ET.fromstring(xml_str), nome)
    except ET.ParseError as e:
        return None, f'XML inválido: {e}'

def ler_xmls_pasta(pasta):
    xmls, erros = [], []
    for arq in os.listdir(pasta):
        if arq.lower().endswith('.xml'):
            dados, erro = ler_xml_arquivo(os.path.join(pasta, arq))
            (xmls if dados else erros).append(dados if dados else f'{arq}: {erro}')
    return xmls, erros

def ler_xmls_receita(xmls_raw):
    xmls, erros = [], []
    for i, xml_str in enumerate(xmls_raw):
        dados, erro = ler_xml_string(xml_str, f'SEFAZ_{i+1}')
        (xmls if dados else erros).append(dados if dados else f'Doc {i+1}: {erro}')
    return xmls, erros


def _res_nfe_to_xml_dict(xml_str):
    """
    Converte um resNFe (resumo SEFAZ) em dict compatível com conferir().
    Campos de detalhe (IPI, ICMS, etc.) ficam como Decimal('0').
    Flag RESUMO=True indica dados parciais.
    """
    try:
        root = ET.fromstring(xml_str)
        def _t(tag):
            el = root.find(f'{NS_NFE}{tag}')
            if el is None:
                el = root.find(tag)
            return (el.text or '').strip() if el is not None else ''
        ch  = _t('chNFe')
        dh  = _t('dhEmi')
        vNF = _t('vNF')
        if len(ch) != 44:
            return None
        zero = Decimal('0')
        return {
            'ARQUIVO'     : f'resNFe_{ch[-8:]}',
            'CHAVE'       : ch,
            'NUMERO'      : str(int(ch[25:34])),
            'SERIE'       : str(int(ch[22:25])),
            'EMISSAO'     : dh[:10] if dh else '',
            'CNPJ_EMIT'   : ch[6:20],
            'FORNECEDOR'  : _t('xNome'),
            'CNPJ_DEST'   : '',
            'DESTINATARIO': '',
            'TOTAL_NF'    : _moeda(vNF) if vNF else zero,
            'TOTAL_PROD'  : zero,
            'TOTAL_IPI'   : zero,
            'TOTAL_FRETE' : zero,
            'TOTAL_ICMS'  : zero,
            'TOTAL_ST'    : zero,
            'DESCONTO'    : zero,
            'CFOP'        : '',
            'RESUMO'      : True,
        }
    except Exception:
        return None


def _info_dest_cnpj(xml_str):
    """Retorna o CNPJ do destinatário de um XML de NF-e (procNFe). '' se falhar."""
    try:
        root = ET.fromstring(xml_str)
        dest_el = root.find(f'.//{NS_NFE}dest')
        if dest_el is None:
            dest_el = root.find('.//dest')
        if dest_el is not None:
            cnpj_el = dest_el.find(f'{NS_NFE}CNPJ')
            if cnpj_el is None:
                cnpj_el = dest_el.find('CNPJ')
            if cnpj_el is not None and cnpj_el.text:
                return ''.join(c for c in cnpj_el.text if c.isdigit())
    except Exception:
        pass
    return ''


def _build_xml_zip(nfs_filtradas, cnpj):
    """Empacota XMLs das NF-e filtradas (procNFe e resNFe) em um ZIP. Retorna (bytes_zip, count)."""
    cache = carregar_cache(cnpj)
    chave_para_xml = {}
    chave_eh_resumo = {}
    for doc in cache:
        schema  = doc.get('schema', '')
        xml_str = doc.get('xml', '')
        is_resumo = 'resNFe' in schema
        is_proc   = 'procNFe' in schema or 'nfeProc' in schema
        if not (is_resumo or is_proc):
            continue
        try:
            root = ET.fromstring(xml_str)
            if is_resumo:
                el = root.find(f'{NS_NFE}chNFe')
                if el is None:
                    el = root.find('chNFe')
                ch = (el.text or '').strip() if el is not None else ''
            else:
                infEl = root.find(f'.//{NS_NFE}infNFe')
                if infEl is None:
                    infEl = root.find('.//infNFe')
                ch = (infEl.get('Id', '') if infEl is not None else '').replace('NFe', '')
            if len(ch) == 44:
                chave_para_xml[ch]  = xml_str
                chave_eh_resumo[ch] = is_resumo
        except Exception:
            continue

    buf   = io.BytesIO()
    count = 0
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for nf in nfs_filtradas:
            ch = (nf.get('CHAVE') or '').strip()
            if not ch or ch not in chave_para_xml:
                continue
            is_res = chave_eh_resumo.get(ch, nf.get('RESUMO', False))
            prefix = 'resumo_' if is_res else ''
            conteudo = chave_para_xml[ch]
            if isinstance(conteudo, str):
                conteudo = conteudo.encode('utf-8')
            zf.writestr(f'{prefix}{ch}.xml', conteudo)
            count += 1
    return buf.getvalue(), count


def _build_danfe_zip(nfs_filtradas, cnpj):
    """
    Gera DANFEs em PDF para as NF-e procNFe filtradas e empacota em ZIP.
    Retorna (bytes_zip, gerados, sem_xml).
    """
    from brazilfiscalreport.danfe import Danfe
    cache = carregar_cache(cnpj)
    chave_para_xml = {}
    for doc in cache:
        schema  = doc.get('schema', '')
        xml_str = doc.get('xml', '')
        if 'procNFe' not in schema and 'nfeProc' not in schema:
            continue
        try:
            root  = ET.fromstring(xml_str)
            infEl = root.find(f'.//{NS_NFE}infNFe')
            if infEl is None:
                infEl = root.find('.//infNFe')
            ch = (infEl.get('Id', '') if infEl is not None else '').replace('NFe', '')
            if len(ch) == 44:
                chave_para_xml[ch] = xml_str
        except Exception:
            continue

    buf     = io.BytesIO()
    gerados = 0
    sem_xml = 0
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for nf in nfs_filtradas:
            ch        = (nf.get('CHAVE') or '').strip()
            is_resumo = nf.get('RESUMO', False)
            if is_resumo or ch not in chave_para_xml:
                sem_xml += 1
                continue
            try:
                pdf_obj   = Danfe(xml=chave_para_xml[ch])
                pdf_bytes = pdf_obj.output()
                zf.writestr(f'{ch}.pdf', bytes(pdf_bytes))
                gerados += 1
            except Exception as exc:
                zf.writestr(f'erro_{ch}.txt', f'Erro ao gerar DANFE para {ch}: {exc}')
    return buf.getvalue(), gerados, sem_xml


def _info_res_nfe(xml_str):
    """Extrai informações básicas de um resNFe (resumo SEFAZ) para exibição."""
    try:
        root = ET.fromstring(xml_str)
        def _t(tag):
            el = root.find(f'{NS_NFE}{tag}')
            if el is None:
                el = root.find(tag)
            return (el.text or '').strip() if el is not None else ''
        ch = _t('chNFe')
        dh = _t('dhEmi')
        return {
            'CHAVE'     : ch,
            'NUMERO'    : str(int(ch[25:34])) if len(ch) == 44 else '?',
            'FORNECEDOR': _t('xNome'),
            'EMISSAO'   : dh[:10] if dh else '',
            'VALOR'     : _fmt_brl(_t('vNF') or '0'),
        }
    except Exception:
        return None

def _parse_nfce(root, nome):
    """Analisa NFC-e (modelo 65) ou SAT (CFe) e retorna dict normalizado."""
    NS = 'http://www.portalfiscal.inf.br/nfe'
    # Detect SAT (CFe)
    is_sat = root.tag.endswith('}CFe') or root.tag == 'CFe'
    if is_sat:
        inf = (root.find(f'{{{NS}}}infCFe') or root.find('infCFe'))
        if inf is None:
            return None, 'infCFe não encontrado'
        chave = (inf.get('Id') or '').replace('CFe', '')
        ide   = inf.find(f'{{{NS}}}ide') or inf.find('ide')
        dest  = inf.find(f'{{{NS}}}dest') or inf.find('dest')
        emit  = inf.find(f'{{{NS}}}emit') or inf.find('emit')
        total = inf.find(f'{{{NS}}}total') or inf.find('total')
        num   = _txt(ide, 'nCFe') if ide is not None else ''
        serie = _txt(ide, 'nserieSAT') if ide is not None else ''
        emissao_raw = _txt(ide, 'dEmi') if ide is not None else ''
        nnf_dest = _txt(dest, 'CNPJ') or _txt(dest, 'CPF') if dest is not None else ''
        nome_dest = _txt(dest, 'xNome') if dest is not None else 'CONSUMIDOR'
        cnpj_emit = _txt(emit, 'CNPJ') if emit is not None else ''
        vCFe      = _txt(total.find(f'{{{NS}}}vCFe') or total.find('vCFe') if total is not None else None, '') or '0'
        vCFe2     = ''
        if total is not None:
            vt = total.find(f'{{{NS}}}vCFe') or total.find('vCFe')
            vCFe2 = vt.text if vt is not None else '0'
        total_nf = _moeda(vCFe2 or '0')
        tipo = 'SAT'
    else:
        # NFC-e: modelo 65, namespace NFe
        inf = root.find(f'{{{NS}}}NFe/{{{NS}}}infNFe') or root.find(f'{{{NS}}}infNFe')
        if inf is None:
            return None, 'infNFe não encontrado'
        chave   = (inf.get('Id') or '').replace('NFe', '')
        ide     = inf.find(f'{{{NS}}}ide')
        dest    = inf.find(f'{{{NS}}}dest')
        emit    = inf.find(f'{{{NS}}}emit')
        total   = inf.find(f'{{{NS}}}total/{{{NS}}}ICMSTot')
        modelo  = _txt(ide, f'{{{NS}}}mod') or _txt(ide, 'mod') if ide is not None else '65'
        if modelo not in ('65',):
            return None, f'Modelo {modelo} não é NFC-e'
        num      = _txt(ide, f'{{{NS}}}nNF') or _txt(ide, 'nNF') if ide is not None else ''
        serie    = _txt(ide, f'{{{NS}}}serie') or _txt(ide, 'serie') if ide is not None else ''
        emissao_raw = _txt(ide, f'{{{NS}}}dhEmi') or _txt(ide, 'dhEmi') if ide is not None else ''
        cnpj_emit = _txt(emit, f'{{{NS}}}CNPJ') if emit is not None else ''
        nnf_dest  = ''
        nome_dest = 'CONSUMIDOR'
        total_nf  = _moeda(_txt(total, f'{{{NS}}}vNF') or _txt(total, 'vNF') or '0') if total is not None else Decimal('0')
        tipo = 'NFC-e'
    try:
        from datetime import datetime
        emissao = datetime.strptime(emissao_raw[:10], '%Y-%m-%d').date() if emissao_raw else None
    except Exception:
        emissao = None
    return {
        'ARQUIVO': nome, 'TIPO': tipo, 'CHAVE': chave,
        'NUMERO': num, 'SERIE': serie, 'EMISSAO': emissao,
        'CNPJ_EMIT': _limpar_cnpj(cnpj_emit),
        'DESTINATARIO': nome_dest, 'CNPJ_DEST': _limpar_cnpj(nnf_dest),
        'TOTAL_NF': total_nf,
    }, None

def ler_cupons_pasta(pasta):
    """Lê todos os XMLs NFC-e/SAT de uma pasta e retorna (cupons, erros)."""
    cupons, erros = [], []
    if not pasta or not os.path.isdir(pasta):
        return cupons, erros
    for arq in sorted(os.listdir(pasta)):
        if not arq.lower().endswith('.xml'):
            continue
        caminho = os.path.join(pasta, arq)
        try:
            tree = ET.parse(caminho)
            root = tree.getroot()
            dados, erro = _parse_nfce(root, arq)
            if dados:
                cupons.append(dados)
            else:
                erros.append(f'{arq}: {erro}')
        except Exception as ex:
            erros.append(f'{arq}: {ex}')
    return cupons, erros


def ler_cupons_banco(con, data_ini, data_fim, cnpj_loja=None):
    """Lê NFC-e e SAT diretamente do banco ERP no período informado.
    Retorna (cupons, erros) com a mesma estrutura que ler_cupons_pasta.
    """
    cupons, erros = [], []
    cur = con.cursor()
    cnpj_limpo = _limpar_cnpj(cnpj_loja) if cnpj_loja else ''
    filtro_loja = f"AND ID_LOJA = (SELECT FIRST 1 ID FROM LOJA WHERE REPLACE(REPLACE(REPLACE(CPFCGC,'.',''),'-',''),'/','') = '{cnpj_limpo}')" if cnpj_limpo else ''
    try:
        cur.execute(
            f"SELECT CHAVE_ACESSO_NFE, NUMERO_NFCE, SERIE_NFCE, DATA, "
            f"TOTAL_LIQUIDO, AUTORIZADO_CANC_NFE, XML, "
            f"COALESCE(STATUS, 'S') AS STATUS "
            f"FROM VENDAS "
            f"WHERE NUMERO_NFCE IS NOT NULL AND CHAVE_ACESSO_NFE IS NOT NULL "
            f"AND DATA BETWEEN ? AND ? "
            f"{filtro_loja} "
            f"ORDER BY DATA, NUMERO_NFCE",
            (data_ini, data_fim),
        )
        for chave, num, serie, data, total, canc_aut, xml_blob, status in cur.fetchall():
            xml_str = ''
            if xml_blob is not None:
                try:
                    raw = xml_blob.read() if hasattr(xml_blob, 'read') else xml_blob
                    xml_str = raw.decode('utf-8', errors='ignore') if isinstance(raw, bytes) else (raw or '')
                    xml_str = xml_str.strip()
                except Exception:
                    xml_str = ''
            cancelado = (status == 'C') or (canc_aut is not None)
            cupons.append({
                'ARQUIVO'     : str(chave or num or ''),
                'TIPO'        : 'NFC-e',
                'CHAVE'       : str(chave or ''),
                'NUMERO'      : str(num or ''),
                'SERIE'       : str(serie or ''),
                'EMISSAO'     : data,
                'CNPJ_EMIT'   : '',
                'DESTINATARIO': 'CONSUMIDOR',
                'CNPJ_DEST'   : '',
                'TOTAL_NF'    : Decimal(str(total or 0)),
                'CANCELADO'   : cancelado,
                'XML'         : xml_str,
            })
    except Exception as e:
        erros.append(f'Erro NFC-e: {e}')
    try:
        cur.execute(
            f"SELECT SAT_CHAVE_ACESSO, DATA, TOTAL_LIQUIDO, SAT_XML, SAT_CHAVE_CANC "
            f"FROM VENDAS "
            f"WHERE SAT_XML IS NOT NULL "
            f"AND (CHAVE_ACESSO_NFE IS NULL AND NUMERO_NFCE IS NULL) "
            f"AND DATA BETWEEN ? AND ? "
            f"{filtro_loja} "
            f"ORDER BY DATA",
            (data_ini, data_fim),
        )
        for chave, data, total, sat_blob, sat_canc in cur.fetchall():
            xml_str = ''
            if sat_blob is not None:
                try:
                    raw = sat_blob.read() if hasattr(sat_blob, 'read') else sat_blob
                    xml_str = raw.decode('utf-8', errors='ignore') if isinstance(raw, bytes) else (raw or '')
                    xml_str = xml_str.strip()
                except Exception:
                    xml_str = ''
            cupons.append({
                'ARQUIVO'     : str(chave or ''),
                'TIPO'        : 'SAT',
                'CHAVE'       : str(chave or ''),
                'NUMERO'      : '',
                'SERIE'       : '',
                'EMISSAO'     : data,
                'CNPJ_EMIT'   : '',
                'DESTINATARIO': 'CONSUMIDOR',
                'CNPJ_DEST'   : '',
                'TOTAL_NF'    : Decimal(str(total or 0)),
                'CANCELADO'   : sat_canc is not None,
                'XML'         : xml_str,
            })
    except Exception as e:
        erros.append(f'Erro SAT: {e}')
    cur.close()
    return cupons, erros


# ─── CONFERÊNCIA ──────────────────────────────────────────────────────────────

def conferir(nfs_erp, nfs_xml):
    xml_por_chave = {}
    xml_por_ident = {}
    xml_usados    = set()

    for xml in nfs_xml:
        if xml['CHAVE']:
            xml_por_chave[xml['CHAVE']] = xml
        ident = (_limpar_cnpj(xml['CNPJ_EMIT']),
                 str(xml['NUMERO']).strip(),
                 (str(xml['SERIE']).strip().lstrip('0') or '0'))
        xml_por_ident[ident] = xml

    def _d(v):
        return v or Decimal('0')

    def _f(v):
        return float(_d(v))

    resultados = []
    campos_val = ['TOTAL_NF', 'TOTAL_PROD', 'TOTAL_IPI',
                  'TOTAL_FRETE', 'TOTAL_ICMS', 'TOTAL_ST', 'DESCONTO']

    for erp in nfs_erp:
        chave_erp = erp['CHAVE'].replace('NFe', '') if erp['CHAVE'] else ''
        xml = xml_por_chave.get(chave_erp) if chave_erp else None
        if xml is None:
            ident = (_limpar_cnpj(erp['CNPJ_FORN']),
                     str(erp['NUMERO'] or '').strip(),
                     (str(erp['SERIE'] or '').strip().lstrip('0') or '0'))
            xml = xml_por_ident.get(ident)

        if xml:
            chave_key = xml.get('CHAVE') or chave_erp
            xml_usados.add(chave_key or id(xml))
            is_resumo = xml.get('RESUMO', False)
            # Para resumos só compara TOTAL_NF (demais campos não existem no resNFe)
            campos_cmp = ['TOTAL_NF'] if is_resumo else campos_val
            divs = [
                f"{c}: ERP={_d(erp.get(c))} XML={_d(xml.get(c))}"
                for c in campos_cmp
                if abs(_d(erp.get(c)) - _d(xml.get(c))) > Decimal('0.05')
            ]
            if is_resumo:
                status = STATUS_RESUMO if not divs else STATUS_DIVERGENCIA
            else:
                status = STATUS_OK if not divs else STATUS_DIVERGENCIA
            r = {
                'Status'          : status,
                'Número'          : erp['NUMERO'],
                'Série'           : erp['SERIE'],
                'Dt. Entrada'     : str(erp['DT_ENTRADA']) if erp['DT_ENTRADA'] else '',
                'Dt. Emissão'     : str(erp['DT_EMISSAO']) if erp['DT_EMISSAO'] else '',
                'Fornecedor ERP'  : erp['FORNECEDOR'],
                'Fornecedor XML'  : xml['FORNECEDOR'],
                'CNPJ'            : erp['CNPJ_FORN'],
                'Loja'            : erp['LOJA'],
                'Total ERP'       : _f(erp['TOTAL_NF']),
                'Total XML'       : _f(xml['TOTAL_NF']),
                'Diferença'       : _f(erp['TOTAL_NF']) - _f(xml['TOTAL_NF']),
                'Prod ERP'        : _f(erp['TOTAL_PROD']),
                'Prod XML'        : _f(xml['TOTAL_PROD']),
                'IPI ERP'         : _f(erp['TOTAL_IPI']),
                'IPI XML'         : _f(xml['TOTAL_IPI']),
                'Frete ERP'       : _f(erp['TOTAL_FRETE']),
                'Frete XML'       : _f(xml['TOTAL_FRETE']),
                'ICMS ERP'        : _f(erp['TOTAL_ICMS']),
                'ICMS XML'        : _f(xml['TOTAL_ICMS']),
                'ST ERP'          : _f(erp['TOTAL_ST']),
                'ST XML'          : _f(xml['TOTAL_ST']),
                'Desc ERP'        : _f(erp['DESCONTO']),
                'Desc XML'        : _f(xml['DESCONTO']),
                'Divergências'    : ' | '.join(divs) if divs else '',
                'Chave NF-e'      : chave_erp or xml['CHAVE'],
                'Arquivo XML'     : xml['ARQUIVO'],
            }
        else:
            r = {
                'Status'          : STATUS_SO_ERP,
                'Número'          : erp['NUMERO'],
                'Série'           : erp['SERIE'],
                'Dt. Entrada'     : str(erp['DT_ENTRADA']) if erp['DT_ENTRADA'] else '',
                'Dt. Emissão'     : str(erp['DT_EMISSAO']) if erp['DT_EMISSAO'] else '',
                'Fornecedor ERP'  : erp['FORNECEDOR'],
                'Fornecedor XML'  : '',
                'CNPJ'            : erp['CNPJ_FORN'],
                'Loja'            : erp['LOJA'],
                'Total ERP'       : _f(erp['TOTAL_NF']),
                'Total XML'       : 0.0,
                'Diferença'       : _f(erp['TOTAL_NF']),
                'Prod ERP'        : _f(erp['TOTAL_PROD']),
                'Prod XML'        : 0.0,
                'IPI ERP'         : _f(erp['TOTAL_IPI']),
                'IPI XML'         : 0.0,
                'Frete ERP'       : _f(erp['TOTAL_FRETE']),
                'Frete XML'       : 0.0,
                'ICMS ERP'        : _f(erp['TOTAL_ICMS']),
                'ICMS XML'        : 0.0,
                'ST ERP'          : _f(erp['TOTAL_ST']),
                'ST XML'          : 0.0,
                'Desc ERP'        : _f(erp['DESCONTO']),
                'Desc XML'        : 0.0,
                'Divergências'    : 'XML não encontrado',
                'Chave NF-e'      : chave_erp,
                'Arquivo XML'     : '',
            }
        resultados.append(r)

    for xml in nfs_xml:
        chave_key = xml.get('CHAVE') or id(xml)
        if chave_key not in xml_usados:
            is_resumo = xml.get('RESUMO', False)
            status_solo = STATUS_SO_XML
            div_msg = 'Não lançada no ERP'
            resultados.append({
                'Status'          : status_solo,
                'Número'          : xml['NUMERO'],
                'Série'           : xml['SERIE'],
                'Dt. Entrada'     : '',
                'Dt. Emissão'     : xml['EMISSAO'],
                'Fornecedor ERP'  : '',
                'Fornecedor XML'  : xml['FORNECEDOR'],
                'CNPJ'            : xml['CNPJ_EMIT'],
                'Loja'            : '',
                'Total ERP'       : 0.0,
                'Total XML'       : _f(xml['TOTAL_NF']),
                'Diferença'       : -_f(xml['TOTAL_NF']),
                'Prod ERP'        : 0.0,
                'Prod XML'        : _f(xml['TOTAL_PROD']),
                'IPI ERP'         : 0.0,
                'IPI XML'         : _f(xml['TOTAL_IPI']),
                'Frete ERP'       : 0.0,
                'Frete XML'       : _f(xml['TOTAL_FRETE']),
                'ICMS ERP'        : 0.0,
                'ICMS XML'        : _f(xml['TOTAL_ICMS']),
                'ST ERP'          : 0.0,
                'ST XML'          : _f(xml['TOTAL_ST']),
                'Desc ERP'        : 0.0,
                'Desc XML'        : _f(xml['DESCONTO']),
                'Divergências'    : div_msg,
                'Chave NF-e'      : xml['CHAVE'],
                'Arquivo XML'     : xml['ARQUIVO'],
            })

    return resultados

# ─── EXCEL ────────────────────────────────────────────────────────────────────

_CORES = {
    STATUS_OK         : 'C6EFCE',
    STATUS_RESUMO     : 'E2EFDA',
    STATUS_DIVERGENCIA: 'FFEB9C',
    STATUS_SO_ERP     : 'FCE4D6',
    STATUS_SO_XML     : 'DAEEF3',
}
_COLUNAS_EXCEL = [
    'Status', 'Número', 'Série', 'Dt. Entrada', 'Dt. Emissão',
    'Fornecedor ERP', 'Fornecedor XML', 'CNPJ', 'Loja',
    'Total ERP', 'Total XML', 'Diferença',
    'Prod ERP', 'Prod XML',
    'IPI ERP', 'IPI XML',
    'Frete ERP', 'Frete XML',
    'ICMS ERP', 'ICMS XML',
    'ST ERP', 'ST XML',
    'Desc ERP', 'Desc XML',
    'Divergências', 'Chave NF-e', 'Arquivo XML',
]

def gerar_excel(resultados):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Conferência NF Entrada'

    hdr_fill = PatternFill('solid', fgColor='1F4E79')
    hdr_font = Font(bold=True, color='FFFFFF')
    for col, titulo in enumerate(_COLUNAS_EXCEL, 1):
        c = ws.cell(row=1, column=col, value=titulo)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 28

    moeda_cols = set(range(10, 25))
    fmt_m = '#,##0.00'

    for lin, r in enumerate(resultados, 2):
        fill = PatternFill('solid', fgColor=_CORES.get(r.get('Status', ''), 'FFFFFF'))
        for col, campo in enumerate(_COLUNAS_EXCEL, 1):
            c = ws.cell(row=lin, column=col, value=r.get(campo, ''))
            c.fill = fill
            c.alignment = Alignment(vertical='center')
            if col in moeda_cols:
                c.number_format = fmt_m

    larguras = {
        1:16, 2:10, 3:6, 4:12, 5:12, 6:30, 7:30, 8:16, 9:22,
        10:14, 11:14, 12:14, 13:13, 14:13, 15:11, 16:11,
        17:11, 18:11, 19:11, 20:11, 21:11, 22:11, 23:11, 24:11,
        25:40, 26:46, 27:28,
    }
    for col, larg in larguras.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = larg

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def gerar_excel_contabilidade(xmls_entrada, xmls_saida, cupons, resultados_ent, resultados_sai, data_ini, data_fim):
    """Gera Excel multi-abas para envio ao contador."""
    _H = PatternFill('solid', fgColor='1F4E79')
    _H_FONT = Font(color='FFFFFF', bold=True)
    def _fmt_hdr(ws, colunas):
        for col_idx, titulo in enumerate(colunas, 1):
            c = ws.cell(row=1, column=col_idx, value=titulo)
            c.fill = _H; c.font = _H_FONT
            c.alignment = Alignment(horizontal='center')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Aba Conferência Entrada ──────────────────────────────────────────────
    if resultados_ent:
        ws = wb.create_sheet('Conferência Entrada')
        cols = list(resultados_ent[0].keys())
        _fmt_hdr(ws, cols)
        COR_STATUS = {STATUS_OK: 'C6EFCE', STATUS_DIVERGENCIA: 'FFEB9C',
                      STATUS_SO_ERP: 'FFCCCC', STATUS_SO_XML: 'DDEBF7'}
        for r in resultados_ent:
            row = [r.get(c) for c in cols]
            ws.append(row)
            cor = COR_STATUS.get(r.get('Status'), 'FFFFFF')
            fill = PatternFill('solid', fgColor=cor)
            for cell in ws[ws.max_row]:
                cell.fill = fill
        ws.freeze_panes = 'A2'; ws.auto_filter.ref = ws.dimensions

    # ── Aba Conferência Saída ────────────────────────────────────────────────
    if resultados_sai:
        ws = wb.create_sheet('Conferência Saída')
        cols = list(resultados_sai[0].keys())
        _fmt_hdr(ws, cols)
        COR_STATUS = {STATUS_OK: 'C6EFCE', STATUS_DIVERGENCIA: 'FFEB9C',
                      STATUS_SO_ERP: 'FFCCCC', STATUS_SO_XML: 'DDEBF7'}
        for r in resultados_sai:
            row = [r.get(c) for c in cols]
            ws.append(row)
            cor = COR_STATUS.get(r.get('Status'), 'FFFFFF')
            fill = PatternFill('solid', fgColor=cor)
            for cell in ws[ws.max_row]:
                cell.fill = fill
        ws.freeze_panes = 'A2'; ws.auto_filter.ref = ws.dimensions

    # ── Aba NF Entrada (lista completa) ─────────────────────────────────────
    if xmls_entrada:
        ws = wb.create_sheet('NF Entrada')
        cols_ent = ['ARQUIVO','CHAVE','NUMERO','SERIE','EMISSAO',
                    'CNPJ_EMIT','FORNECEDOR','TOTAL_NF','TOTAL_PROD',
                    'TOTAL_IPI','TOTAL_FRETE','TOTAL_ICMS','TOTAL_ST','DESCONTO']
        _fmt_hdr(ws, cols_ent)
        for r in xmls_entrada:
            ws.append([str(r.get(c, '')) for c in cols_ent])
        ws.freeze_panes = 'A2'; ws.auto_filter.ref = ws.dimensions

    # ── Aba NF Saída (lista completa) ───────────────────────────────────────
    if xmls_saida:
        ws = wb.create_sheet('NF Saída')
        cols_sai = ['ARQUIVO','CHAVE','NUMERO','SERIE','EMISSAO',
                    'CNPJ_EMIT','FORNECEDOR','CNPJ_DEST','DESTINATARIO',
                    'TOTAL_NF','TOTAL_PROD','TOTAL_IPI','TOTAL_FRETE',
                    'TOTAL_ICMS','TOTAL_ST','DESCONTO']
        _fmt_hdr(ws, cols_sai)
        for r in xmls_saida:
            ws.append([str(r.get(c, '')) for c in cols_sai])
        ws.freeze_panes = 'A2'; ws.auto_filter.ref = ws.dimensions

    # ── Aba Cupons NFC-e/SAT ─────────────────────────────────────────────────
    if cupons:
        ws = wb.create_sheet('Cupons NFC-e SAT')
        cols_cup = ['ARQUIVO','TIPO','CHAVE','NUMERO','SERIE','EMISSAO',
                    'CNPJ_EMIT','DESTINATARIO','CNPJ_DEST','TOTAL_NF']
        _fmt_hdr(ws, cols_cup)
        for r in cupons:
            ws.append([str(r.get(c, '')) for c in cols_cup])
        ws.freeze_panes = 'A2'; ws.auto_filter.ref = ws.dimensions

    # ── Aba Resumo ───────────────────────────────────────────────────────────
    ws = wb.create_sheet('Resumo')
    ws.sheet_properties.tabColor = '1F4E79'
    resumo_data = [
        ['Período', f'{data_ini} a {data_fim}'],
        [],
        ['Tipo', 'Qtde', 'Valor Total (R$)'],
        ['NF Entrada (SEFAZ)',
         len(xmls_entrada),
         float(sum(r.get('TOTAL_NF', 0) for r in xmls_entrada))],
        ['NF Saída (SEFAZ)',
         len(xmls_saida),
         float(sum(r.get('TOTAL_NF', 0) for r in xmls_saida))],
        ['Cupons NFC-e/SAT',
         len(cupons),
         float(sum(r.get('TOTAL_NF', 0) for r in cupons))],
    ]
    for row in resumo_data:
        ws.append(row)
    # Header styling on row 3
    for cell in ws[3]:
        if cell.value:
            cell.fill = _H; cell.font = _H_FONT
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def _slug_loja(nome):
    """Converte nome de loja para slug seguro para uso em nome de arquivo."""
    import unicodedata, re as _re2
    nfkd = unicodedata.normalize('NFKD', str(nome))
    s = ''.join(c for c in nfkd if not unicodedata.combining(c))
    s = _re2.sub(r'[^\w]', '_', s.upper())
    s = _re2.sub(r'_+', '_', s).strip('_')
    return s


def _data_xml(xml_str):
    """Extrai data de emissão (dhEmi ou dEmi) de um XML NF-e. Retorna 'YYYY-MM-DD' ou ''."""
    try:
        root = ET.fromstring(xml_str)
        for tag in ('dhEmi', 'dEmi'):
            el = root.find(f'.//{NS_NFE}{tag}')
            if el is None:
                el = root.find(f'.//{tag}')
            if el is not None and el.text:
                return str(el.text)[:10]
    except Exception:
        pass
    return ''


def gerar_zip_loja_bytes(loja_cfg, db_cfg_central, cont_ini, cont_fim):
    """Gera ZIP com XMLs NF-e (Entrada/, Saida/) do cache SEFAZ + Cupons/ do ERP
    para a loja e período informados. Retorna bytes ou None se sem dados."""
    cnpj   = _limpar_cnpj(loja_cfg.get('cnpj', ''))
    loja_id = loja_cfg.get('id')
    d0, d1 = str(cont_ini), str(cont_fim)

    # ── Entrada: cache SEFAZ ──────────────────────────────────────────────────
    docs_ent = []
    if cnpj and SEFAZ_OK:
        docs_raw = carregar_cache(cnpj)
        docs_ent = [x for x in filtrar_nfes(docs_raw, cnpj)
                    if (lambda d: not d or d0 <= d <= d1)(_data_xml(x))]

    # ── Determina conexão ERP da loja ─────────────────────────────────────────
    if loja_cfg.get('db_host'):
        loja_db = {
            'host'    : loja_cfg['db_host'],
            'port'    : loja_cfg.get('db_port', 3050),
            'database': loja_cfg['db_database'],
            'user'    : loja_cfg.get('db_user', 'SYSDBA'),
            'password': loja_cfg.get('db_password', 'masterkey'),
        }
    elif loja_id == 1:
        loja_db = db_cfg_central
    else:
        loja_db = None

    # ── Saída: busca diretamente do ERP (tabela NFISCAL) ─────────────────────
    docs_sai = []
    if loja_db:
        try:
            _con_s = conectar_erp(loja_db)
            _rows_s = buscar_nfs_saida_erp(_con_s, cont_ini, cont_fim, loja_id=loja_id)
            _con_s.close()
            docs_sai = [r['XML'] for r in _rows_s if r.get('XML')]
        except Exception:
            pass

    # ── Cupons: todas as lojas EXCETO loja 1 (FRAMODAS não tem NFC-e/SAT) ────
    cupons_xml = []
    if loja_db and loja_id != 1:
        try:
            _con_c = conectar_erp(loja_db)
            _cups, _ = ler_cupons_banco(_con_c, cont_ini, cont_fim, cnpj or None)
            _con_c.close()
            cupons_xml = [c['XML'] for c in _cups if c.get('XML')]
        except Exception:
            pass

    if not docs_ent and not docs_sai and not cupons_xml:
        return None

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for i, xml_str in enumerate(docs_ent, 1):
            zf.writestr(f'Entrada/nfe_entrada_{i:04d}.xml', xml_str)
        for i, xml_str in enumerate(docs_sai, 1):
            chave_s = ''
            try:
                _rs = ET.fromstring(xml_str)
                for _tg in ('chNFe', 'chAcesso'):
                    _el = _rs.find(f'.//{NS_NFE}{_tg}') or _rs.find(f'.//{_tg}')
                    if _el is not None and _el.text:
                        chave_s = _el.text.strip()[:44]
                        break
            except Exception:
                pass
            zf.writestr(f'Saida/{chave_s or f"nfe_saida_{i:04d}"}.xml', xml_str)
        for idx_c, cup in enumerate(cupons_xml, 1):
            if cup:
                try:
                    _r = ET.fromstring(cup)
                    _ch = ''
                    for _tag in ('chNFe', 'chAcesso', 'chaveConsulta'):
                        _el = _r.find(f'.//{NS_NFE}{_tag}') or _r.find(f'.//{_tag}')
                        if _el is not None and _el.text:
                            _ch = _el.text.strip()[:44]
                            break
                    _nome_cup = f'Cupons/{_ch or f"cupon_{idx_c:04d}"}.xml'
                except Exception:
                    _nome_cup = f'Cupons/cupon_{idx_c:04d}.xml'
                zf.writestr(_nome_cup, cup)
    buf.seek(0)
    return buf.getvalue()


def enviar_email_contabilidade(zips_por_loja, remetente, senha_app, destinatario, mes_ref, brevo_api_key=''):
    """Envia e-mail com ZIPs. Se brevo_api_key, usa Brevo API (rastreamento); senão Gmail SMTP."""
    import base64 as _b64

    corpo_txt = (
        f'Prezado(a) Contador(a),\n\n'
        f'Seguem em anexo os arquivos XML fiscais referentes ao período {mes_ref}, '
        f'organizados por loja (pastas: Entrada/, Saida/, Cupons/).\n\n'
        f'Atenciosamente,\n'
        f'Framodas'
    )

    if brevo_api_key:
        import requests as _req
        attachments = []
        for nome_arq, dados in zips_por_loja:
            attachments.append({
                'name': nome_arq,
                'content': _b64.b64encode(dados).decode(),
            })
        # Normalizar destinatários para lista
        if isinstance(destinatario, str):
            _dests = [e.strip() for e in destinatario.split(',') if e.strip()]
        else:
            _dests = list(destinatario)
        _dests_extras = ['arthur.mataveli@controltech.com.br', 'blink_rh@lojasblink.com.br']
        for _e in _dests_extras:
            if _e not in _dests:
                _dests.append(_e)

        payload = {
            'sender': {'email': remetente, 'name': 'Framodas CPD'},
            'to': [{'email': e} for e in _dests],
            'subject': f'XMLs Fiscais — {mes_ref}',
            'textContent': corpo_txt,
        }
        if attachments:
            payload['attachment'] = attachments
        resp = _req.post(
            'https://api.brevo.com/v3/smtp/email',
            json=payload,
            headers={'api-key': brevo_api_key, 'Content-Type': 'application/json'},
            timeout=120,
        )
        if not resp.ok:
            try:
                _detail = resp.json().get('message', resp.text)
            except Exception:
                _detail = resp.text
            raise RuntimeError(f'Brevo API {resp.status_code}: {_detail}')
    else:
        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.base import MIMEBase
        from email.mime.text import MIMEText
        from email import encoders

        msg = MIMEMultipart()
        msg['From']    = remetente
        msg['To']      = ', '.join(_dests)
        msg['Subject'] = f'XMLs Fiscais — {mes_ref}'
        msg.attach(MIMEText(corpo_txt, 'plain', 'utf-8'))

        for nome_arq, dados in zips_por_loja:
            part = MIMEBase('application', 'zip')
            part.set_payload(dados)
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{nome_arq}"')
            msg.attach(part)

        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(remetente, senha_app)
        server.sendmail(remetente, _dests, msg.as_string())
        server.quit()


def consultar_status_emails_brevo(brevo_api_key, dias=30):
    """Consulta eventos de e-mail transacional no Brevo (delivered, opened, clicked, etc.)."""
    import requests as _req
    from datetime import date, timedelta

    _end = date.today()
    _start = _end - timedelta(days=dias)
    eventos_raw = []
    offset = 0
    while True:
        resp = _req.get(
            'https://api.brevo.com/v3/smtp/statistics/events',
            headers={'api-key': brevo_api_key, 'Accept': 'application/json'},
            params={
                'startDate': _start.isoformat(),
                'endDate': _end.isoformat(),
                'limit': 2500,
                'offset': offset,
                'sort': 'desc',
            },
            timeout=30,
        )
        if not resp.ok:
            break
        data = resp.json()
        evts = data.get('events', [])
        if not evts:
            break
        eventos_raw.extend(evts)
        offset += len(evts)
        if len(evts) < 2500:
            break
    # Agrupar por messageId
    msgs = {}
    for ev in eventos_raw:
        mid = ev.get('messageId', '')
        if mid not in msgs:
            msgs[mid] = {
                'email': ev.get('email', ''),
                'subject': ev.get('subject', ''),
                'date': ev.get('date', ''),
                'events': [],
            }
        msgs[mid]['events'].append(ev.get('event', ''))
    # Determinar status final de cada mensagem
    resultado = []
    _STATUS_PT = {
        'opened': '✅ Aberto',
        'clicks': '🔗 Clicado',
        'delivered': '📨 Entregue',
        'requests': '⏳ Enviado',
        'deferred': '⚠️ Adiado',
        'softBounces': '⚠️ Bounce temporário',
        'hardBounces': '❌ Bounce permanente',
        'blocked': '🚫 Bloqueado',
        'spam': '🚫 Spam',
        'invalid': '❌ Inválido',
        'unsubscribed': '🚫 Descadastrado',
    }
    _PRIORIDADE = ['opened', 'clicks', 'delivered', 'requests', 'deferred',
                   'softBounces', 'hardBounces', 'blocked', 'spam', 'invalid', 'unsubscribed']
    for mid, info in msgs.items():
        evts = set(info['events'])
        status_key = 'requests'
        for p in _PRIORIDADE:
            if p in evts:
                status_key = p
                break
        resultado.append({
            'Data': info['date'][:16].replace('T', ' '),
            'Destinatário': info['email'],
            'Assunto': info['subject'],
            'Status': _STATUS_PT.get(status_key, status_key),
        })
    return resultado


def gerar_zip_xmls(docs_entrada_raw, docs_saida_raw, pasta_cupons):
    """Empacota XMLs de NF entrada/saída + cupons numa estrutura ZIP."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for i, xml_str in enumerate(docs_entrada_raw, 1):
            zf.writestr(f'Entrada/nfe_entrada_{i:04d}.xml', xml_str)
        for i, xml_str in enumerate(docs_saida_raw, 1):
            zf.writestr(f'Saida/nfe_saida_{i:04d}.xml', xml_str)
        if pasta_cupons and os.path.isdir(pasta_cupons):
            for arq in sorted(os.listdir(pasta_cupons)):
                if arq.lower().endswith('.xml'):
                    caminho = os.path.join(pasta_cupons, arq)
                    try:
                        zf.write(caminho, f'Cupons/{arq}')
                    except Exception:
                        pass
    buf.seek(0)
    return buf.getvalue()

# ─── STREAMLIT APP ────────────────────────────────────────────────────────────

st.set_page_config(
    page_title='Gestão Fiscal — Framodas',
    page_icon='🧾',
    layout='wide',
    menu_items={
        'Get help': None,
        'Report a bug': None,
        'About': '**Gestão Fiscal Framodas** — Conferência de NF-e vs Receita Federal',
    },
)

# Session state
_DEFAULTS = {
    'nfs_xml_mem'         : [],
    'nfs_xml_saida_mem'   : [],
    'nfs_erp_saida_lista' : [],
    'cupons_mem'          : [],
    'resultados'          : [],
    'resultados_saida'    : [],
    'sefaz_logs'       : [],
    'db_cfg'           : None,
    'cert_cfg'         : None,
    'data_ini'         : date.today().replace(day=1),
    'data_fim'         : date.today(),
}
for k, v in _DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v

if st.session_state.db_cfg is None:
    st.session_state.db_cfg = load_db_config()
if st.session_state.cert_cfg is None:
    st.session_state.cert_cfg = load_cert_config()

# ── Seletor de loja (sidebar) ─────────────────────────────────────────────────
_lojas_cfg = load_lojas_config()
_loja_ativa = None
if _lojas_cfg:
    with st.sidebar:
        st.markdown('### 🏪 Loja ativa')
        _loja_opts = [f"Loja {l['id']:02d} — {l['nome']}" for l in _lojas_cfg]
        _loja_sel  = st.selectbox('Loja', _loja_opts, key='sel_loja_ativa',
                                  label_visibility='collapsed')
        _loja_idx  = _loja_opts.index(_loja_sel)
        _loja_sb   = _lojas_cfg[_loja_idx]
        _pfx_abs   = os.path.join(BASE_DIR, _loja_sb['pfx_rel'].replace('/', os.sep))
        _pfx_ok    = os.path.exists(_pfx_abs)
        st.caption(f'CNPJ: `{_loja_sb["cnpj"]}`')
        st.caption('🔐 Cert: ' + ('✅ encontrado' if _pfx_ok else '❌ não encontrado'))
        if _loja_sb.get('tabela_entrada') == 'NFISCAL':
            st.caption('📄 Entrada via: `NFISCAL`')
        else:
            st.caption('📄 Entrada via: `ENTRADA_NF`')
        if _loja_sb.get('obs'):
            st.caption(f'⚠️ {_loja_sb["obs"]}')

        st.sidebar.markdown('---')
        st.sidebar.markdown(
            '<div style="text-align:center;padding:6px 0 2px 0">'
            '<span style="font-size:0.72em;color:#888">Desenvolvido por</span><br>'
            '<span style="font-size:0.85em;font-weight:600;color:#1a56db">Fabrício Zamprogno</span>'
            '</div>',
            unsafe_allow_html=True,
        )

    # Quando loja muda, atualiza cert_cfg e recarrega cache SEFAZ
    if st.session_state.get('_loja_id_ativa') != _loja_sb['id']:
        st.session_state['_loja_id_ativa'] = _loja_sb['id']
        _ccfg_old = st.session_state.cert_cfg or {}
        st.session_state.cert_cfg = {
            'pfx_path'  : _pfx_abs,
            'senha'     : _loja_sb['senha_pfx'],
            'cnpj'      : _loja_sb['cnpj'],
            'uf'        : _loja_sb.get('uf', 32),
            'pasta_nfce': _ccfg_old.get('pasta_nfce', ''),
        }
        # Limpa resultados e recarrega cache SEFAZ da nova loja
        st.session_state.resultados           = []
        st.session_state.resultados_saida     = []
        st.session_state.nfs_erp_saida_lista  = []
        _cnpj_nova = _loja_sb['cnpj']
        _docs_nova = carregar_cache(_cnpj_nova) if SEFAZ_OK else []
        _xmls_nova = []
        if _docs_nova:
            _xmls_nova, _ = ler_xmls_receita(filtrar_nfes(_docs_nova, _cnpj_nova))
            _chp_nova = {x.get('CHAVE', '') for x in _xmls_nova}
            for _dl in _docs_nova:
                if 'resNFe' in _dl.get('schema', ''):
                    _rd = _res_nfe_to_xml_dict(_dl['xml'])
                    if _rd and _rd.get('CHAVE', '') not in _chp_nova:
                        _xmls_nova.append(_rd)
        # Se a loja tem cnpj_matriz, também carrega NF-e do cache da MATRIZ
        _cnpj_mtz = _loja_sb.get('cnpj_matriz')
        if _cnpj_mtz and SEFAZ_OK:
            _docs_mtz = carregar_cache(_cnpj_mtz)
            if _docs_mtz:
                _xmls_mtz, _ = ler_xmls_receita(filtrar_nfes(_docs_mtz, _cnpj_mtz))
                _chp_nova = {x.get('CHAVE', '') for x in _xmls_nova}
                for _xm in _xmls_mtz:
                    if _xm.get('CHAVE', '') and _xm['CHAVE'] not in _chp_nova:
                        _xmls_nova.append(_xm)
                        _chp_nova.add(_xm['CHAVE'])
                # resNFe da matriz também
                for _dl in _docs_mtz:
                    if 'resNFe' in _dl.get('schema', ''):
                        _rd = _res_nfe_to_xml_dict(_dl['xml'])
                        if _rd and _rd.get('CHAVE', '') not in _chp_nova:
                            _xmls_nova.append(_rd)
                            _chp_nova.add(_rd['CHAVE'])
        st.session_state.nfs_xml_mem = _xmls_nova
        _xmlss_nova = []
        if _docs_nova:
            _xmlss_nova, _ = ler_xmls_receita(filtrar_nfes_saida(_docs_nova, _cnpj_nova))
        st.session_state.nfs_xml_saida_mem = _xmlss_nova
        st.rerun()

    # Deriva _loja_ativa para uso nas tabs
    for _l in _lojas_cfg:
        if _l['id'] == st.session_state.get('_loja_id_ativa'):
            _loja_ativa = _l
            break
    if _loja_ativa is None:
        _loja_ativa = _lojas_cfg[0]
else:
    # Fallback: sem lojas_config, carrega cache pelo cert_config.json
    if SEFAZ_OK and not st.session_state.nfs_xml_mem:
        cnpj0 = _limpar_cnpj(st.session_state.cert_cfg.get('cnpj', ''))
        if cnpj0:
            docs0 = carregar_cache(cnpj0)
            if docs0:
                xmls0, _ = ler_xmls_receita(filtrar_nfes(docs0, cnpj0))
                _chaves_proc0 = {x.get('CHAVE', '') for x in xmls0}
                for _d0 in docs0:
                    if 'resNFe' in _d0.get('schema', ''):
                        _rd = _res_nfe_to_xml_dict(_d0['xml'])
                        if _rd and _rd.get('CHAVE', '') not in _chaves_proc0:
                            xmls0.append(_rd)
                if xmls0:
                    st.session_state.nfs_xml_mem = xmls0
                xmls0s, _ = ler_xmls_receita(filtrar_nfes_saida(docs0, cnpj0))
                if xmls0s:
                    st.session_state.nfs_xml_saida_mem = xmls0s

# ─── AUTENTICAÇÃO ────────────────────────────────────────────────────────────
def _load_login_config():
    """Carrega credenciais de login_config.json"""
    LOGIN_CONFIG = os.path.join(BASE_DIR, 'login_config.json')
    try:
        with open(LOGIN_CONFIG, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        st.error(f'Erro ao carregar login_config.json: {e}')
        return None

def _autenticar():
    """Tela de login profissional"""
    _login_cfg = _load_login_config()
    if not _login_cfg:
        st.error('Configuração de login não encontrada.')
        st.stop()
    
    st.set_page_config(page_title=_login_cfg.get('titulo_app', 'FRAMODAS'), layout='centered')
    
    # CSS customizado
    st.markdown('''
        <style>
        #MainMenu {visibility:hidden}
        footer {visibility:hidden}
        header {visibility:hidden}
        .login-container {
            max-width: 400px;
            margin: 0 auto;
            padding: 40px 20px;
        }
        .login-title {
            text-align: center;
            color: #1F4E79;
            font-size: 1.8em;
            font-weight: 700;
            margin-bottom: 8px;
        }
        .login-subtitle {
            text-align: center;
            color: #666;
            font-size: 0.95em;
            margin-bottom: 32px;
        }
        </style>
    ''', unsafe_allow_html=True)
    
    # Container
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown('<div class="login-title">🔒 FRAMODAS</div>', unsafe_allow_html=True)
        st.markdown('<div class="login-subtitle">Conferência NF Entrada</div>', unsafe_allow_html=True)
        
        st.markdown('---')
        
        _user = st.text_input('👤 Usuário', key='login_user', placeholder='Insira seu usuário')
        _pass = st.text_input('🔑 Senha', type='password', key='login_pass', placeholder='Insira sua senha')
        
        st.markdown('---')
        
        if st.button('🔓 Entrar', use_container_width=True, type='primary'):
            if _user == _login_cfg.get('usuario') and _pass == _login_cfg.get('senha'):
                st.session_state['_autenticado'] = True
                st.rerun()
            else:
                st.error('❌ Usuário ou senha incorretos!')
        
        st.markdown('---')
        st.caption(f'📧 Suporte: {_login_cfg.get("email_suporte", "ti@framodas.com.br")}')

# Verifica autenticação
if 'st_session_state_data_ini' not in st.session_state or not hasattr(st.session_state, '_autenticado'):
    st.session_state._autenticado = False

if not st.session_state._autenticado:
    _autenticar()
    st.stop()

# ─── INTERFACE PRINCIPAL (após autenticação) ───────────────────────────────────

st.set_page_config(page_title=_load_login_config().get('titulo_app', 'FRAMODAS'), layout='wide')

# Cabeçalho
st.markdown(
    '<style>'
    '#MainMenu {visibility:hidden}'
    'footer {visibility:hidden}'
    'header {visibility:hidden}'
    '</style>',
    unsafe_allow_html=True,
)
st.markdown(
    '<div style="background:#1F4E79;padding:10px 18px;border-radius:6px;margin-bottom:12px">'
    '<span style="color:#fff;font-size:1.35em;font-weight:700">📋 Gestão Fiscal — ERP vs Receita Federal</span> <span style="color:#FFD700;font-size:0.9em;margin-left:14px">⚙️ ' + _VERSAO_MARCA + '</span>'
    '</div>',
    unsafe_allow_html=True,
)

tab_sefaz, tab_conf, tab_entrada, tab_saida, tab_cupons, tab_contab, tab_cfg = st.tabs([
    '🏛️ Receita Federal (SEFAZ)',
    '📋 Conf. NF Entrada',
    '📥 NF de Entrada',
    '📤 NF de Saída',
    '🧾 Cupons NFC-e/SAT',
    '📦 Contabilidade',
    '⚙️ Configurações',
])

# ══════════════════════════════════════════════════════════════════════════════
# TAB: CONFIGURAÇÕES
# ══════════════════════════════════════════════════════════════════════════════
with tab_cfg:
    # ── Proteção por senha ────────────────────────────────────────────────────
    _CFG_SENHA_CORRETA = 'admin123'
    if not st.session_state.get('_cfg_autenticado'):
        st.markdown('### 🔒 Área restrita')
        _senha_dig = st.text_input('Senha de acesso', type='password',
                                   key='inp_cfg_senha', placeholder='Digite a senha...')
        if st.button('Entrar', key='btn_cfg_entrar', type='primary'):
            if _senha_dig == _CFG_SENHA_CORRETA:
                st.session_state['_cfg_autenticado'] = True
                st.rerun()
            else:
                st.error('Senha incorreta.')
    else:
        if st.button('🔒 Sair das Configurações', key='btn_cfg_sair'):
            st.session_state['_cfg_autenticado'] = False
            st.rerun()

        st.subheader('Conexão com o Banco de Dados (Firebird)')
        st.info(
            '💡 Cada loja/computador pode ter um IP e caminho de banco diferentes. '
            'Configure aqui e salve — as configurações ficam gravadas nesta máquina.',
        )
    
        # ── Auto-leitura do VisualControl.ini ─────────────────────────────────────
        vc_ini = ler_config_visualcontrol()
        if vc_ini:
            st.success(f'✔ VisualControl.ini detectado: `{vc_ini["_ini_path"]}`')
            if st.button('📁 Preencher automaticamente do ERP (VisualControl.ini)', key='btn_autoread_vc'):
                novo_cfg = {k: v for k, v in vc_ini.items() if k != '_ini_path'}
                _save_json(DB_CONFIG, novo_cfg)
                st.session_state.db_cfg = novo_cfg
                st.rerun()
        else:
            st.info('VisualControl.ini não encontrado nos caminhos padrão. Preencha manualmente.')
    
        cfg = st.session_state.db_cfg
        col1, col2 = st.columns([3, 1])
        with col1:
            h_host = st.text_input('IP / Host do servidor ERP', value=cfg['host'])
            h_db   = st.text_input(
                'Caminho do banco de dados (.FDB)',
                value=cfg['database'],
                help=r'Ex: C:\Program Files (x86)\VisualControl\Database\ERP_VCONTROL_BLINK.FDB',
            )
            h_user = st.text_input('Usuário', value=cfg['user'])
        with col2:
            h_port = st.number_input('Porta', value=int(cfg['port']),
                                     min_value=1, max_value=65535, step=1)
            h_pass = st.text_input('Senha', value=cfg['password'], type='password')
    
        bc1, bc2, _ = st.columns([1, 1, 4])
        with bc1:
            if st.button('💾 Salvar', type='primary', key='btn_salvar_cfg'):
                novo = {'host': h_host, 'port': int(h_port),
                        'database': h_db, 'user': h_user, 'password': h_pass}
                _save_json(DB_CONFIG, novo)
                st.session_state.db_cfg = novo
                st.success('Configuração salva!')
        with bc2:
            if st.button('🔗 Testar conexão', key='btn_test_cfg'):
                with st.spinner('Conectando...'):
                    try:
                        test = {'host': h_host, 'port': int(h_port),
                                'database': h_db, 'user': h_user, 'password': h_pass}
                        con = conectar_erp(test)
                        con.close()
                        st.success('✔ Conexão bem-sucedida!')
                    except Exception as e:
                        st.error(f'Erro: {e}')
    
        # ── Banco de dados por loja (lojas com ERP local) ─────────────────────────
        st.markdown('---')
        st.subheader('🏪 Banco de dados por loja')
        st.info(
            'As lojas que possuem o VisualControl instalado **localmente** (07, 08, 11, 13, 16, 17) '
            'precisam ter seu IP e caminho do banco configurados aqui. '
            'Quando configuradas, o app conecta diretamente no banco delas para buscar NF de saída e cupons.',
            icon='💡',
        )
        _lojas_locais = [l for l in load_lojas_config() if l['id'] != 1]
        if not _lojas_locais:
            st.warning('Nenhuma loja encontrada no lojas_config.json.')
        else:
            for _ll in _lojas_locais:
                with st.expander(f"Loja {_ll['id']:02d} — {_ll['nome']}  (CNPJ: {_ll['cnpj']})", expanded=bool(_ll.get('db_host'))):
                    _lk = f"lj{_ll['id']}"
                    _lc1, _lc2 = st.columns([3, 1])
                    with _lc1:
                        _v_host = st.text_input('IP / Host', value=_ll.get('db_host', ''),
                                                key=f'ldb_host_{_lk}',
                                                placeholder='ex: 192.168.15.9')
                        _v_db   = st.text_input('Caminho do banco (.FDB)', value=_ll.get('db_database', ''),
                                                key=f'ldb_db_{_lk}',
                                                placeholder=r'ex: C:\Program Files (x86)\VisualControl\Database\campogrande.FDB')
                        _v_user = st.text_input('Usuário', value=_ll.get('db_user', 'SYSDBA'),
                                                key=f'ldb_user_{_lk}')
                    with _lc2:
                        _v_port = st.number_input('Porta', value=int(_ll.get('db_port', 3050)),
                                                  min_value=1, max_value=65535, step=1,
                                                  key=f'ldb_port_{_lk}')
                        _v_pass = st.text_input('Senha', value=_ll.get('db_password', ''),
                                                type='password', key=f'ldb_pass_{_lk}')
                    _lb1, _lb2, _ = st.columns([1, 1, 4])
                    with _lb1:
                        if st.button('💾 Salvar', key=f'btn_salvar_ldb_{_lk}', type='primary'):
                            _lojas_all = load_lojas_config()
                            for _la in _lojas_all:
                                if _la['id'] == _ll['id']:
                                    if _v_host.strip():
                                        _la['db_host']     = _v_host.strip()
                                        _la['db_port']     = int(_v_port)
                                        _la['db_database'] = _v_db.strip()
                                        _la['db_user']     = _v_user.strip()
                                        _la['db_password'] = _v_pass
                                    else:
                                        # Limpa se host foi apagado
                                        for _fk in ('db_host','db_port','db_database','db_user','db_password'):
                                            _la.pop(_fk, None)
                                    break
                            _save_json(LOJAS_CONFIG, _lojas_all)
                            st.success(f'Salvo para Loja {_ll["id"]}!')
                            st.rerun()
                    with _lb2:
                        if st.button('🔗 Testar', key=f'btn_test_ldb_{_lk}'):
                            if not _v_host.strip():
                                st.error('Informe o IP.')
                            elif not _v_db.strip():
                                st.error('Informe o caminho do banco.')
                            else:
                                with st.spinner('Conectando...'):
                                    try:
                                        _test_cfg = {'host': _v_host.strip(),
                                                     'port': int(_v_port),
                                                     'database': _v_db.strip(),
                                                     'user': _v_user.strip(),
                                                     'password': _v_pass}
                                        _tc = conectar_erp(_test_cfg)
                                        _tc.close()
                                        st.success('✔ Conexão bem-sucedida!')
                                    except Exception as _te:
                                        st.error(f'Erro: {_te}')
    
        # ── E-mail para Contabilidade ──────────────────────────────────────────────
        st.markdown('---')
        st.subheader('📧 E-mail para Contabilidade')
        st.info(
            '**Recomendado:** use o **Brevo** (gratuito) para ter rastreamento de abertura. '
            'Cadastre em [brevo.com](https://brevo.com), vá em **SMTP & API → API Keys** e cole a chave abaixo. '
            'Se deixar a chave vazia, envia pelo Gmail SMTP normalmente.',
            icon='💡',
        )
        _ecfg_s = load_email_config()
        _es1, _es2 = st.columns(2)
        with _es1:
            _e_rem  = st.text_input('Remetente (Gmail da empresa)', value=_ecfg_s.get('remetente', 'blinkcpd@gmail.com'), key='cfg_email_rem')
            _e_dest = st.text_input('Destinatário (e-mail do contador)', value=_ecfg_s.get('destinatario', ''), key='cfg_email_dest',
                                     placeholder='contador@escritorio.com.br')
        with _es2:
            _e_brevo = st.text_input('🔑 Brevo API Key (recomendado)',
                                      value=_ecfg_s.get('brevo_api_key', ''),
                                      type='password', key='cfg_email_brevo',
                                      placeholder='xkeysib-...',
                                      help='Chave de API do Brevo — habilita rastreamento de abertura.')
            _e_senha = st.text_input('Senha de App Gmail (fallback se sem Brevo)',
                                      value=_ecfg_s.get('senha_app', ''),
                                      type='password', key='cfg_email_senha',
                                      help='Usado somente se a API Key do Brevo estiver vazia.')

        if _ecfg_s.get('brevo_api_key'):
            st.success('✅ Brevo configurado — e-mails enviados com rastreamento de abertura.')
        else:
            st.warning('⚠️ Sem Brevo API Key — usando Gmail SMTP (sem rastreamento).', icon='⚠️')

        _eb1, _eb2, _ = st.columns([1, 1, 4])
        with _eb1:
            if st.button('💾 Salvar e-mail', key='btn_salvar_email_cfg', type='primary'):
                _save_json(EMAIL_CONFIG, {
                    'remetente'    : _e_rem.strip(),
                    'senha_app'    : _e_senha,
                    'destinatario' : _e_dest.strip(),
                    'brevo_api_key': _e_brevo.strip(),
                })
                st.success('✔ Configuração de e-mail salva!')
        with _eb2:
            if st.button('📧 Enviar e-mail de teste', key='btn_test_email_cfg'):
                _ecfg_t = load_email_config()
                if not _ecfg_t.get('remetente') or not _ecfg_t.get('destinatario'):
                    st.error('Salve as configurações de e-mail primeiro.')
                elif not _ecfg_t.get('brevo_api_key') and not _ecfg_t.get('senha_app'):
                    st.error('Informe a Brevo API Key ou a Senha de App Gmail.')
                else:
                    with st.spinner('Enviando e-mail de teste...'):
                        try:
                            enviar_email_contabilidade(
                                [],
                                _ecfg_t['remetente'],
                                _ecfg_t.get('senha_app', ''),
                                _ecfg_t['destinatario'],
                                'Teste de configuração',
                                brevo_api_key=_ecfg_t.get('brevo_api_key', ''),
                            )
                            _metodo = 'Brevo' if _ecfg_t.get('brevo_api_key') else 'Gmail SMTP'
                            st.success(f'✔ E-mail de teste enviado via {_metodo} para {_ecfg_t["destinatario"]}!')
                        except Exception as _et:
                            st.error(f'Erro ao enviar: {_et}')
    
    # ══════════════════════════════════════════════════════════════════════════════
# TAB: RECEITA FEDERAL (SEFAZ)
# ══════════════════════════════════════════════════════════════════════════════
with tab_sefaz:
    if not SEFAZ_OK:
        st.error('Módulo sefaz_distdfe.py não encontrado na pasta do programa.')
        st.stop()

    st.subheader('Buscar XMLs direto na Receita Federal (DistDFeInt)')

    # ── Certificado A1 instalado no Windows ──────────────────────────────────
    with st.expander('🪟 Usar certificado A1 instalado no Windows (recomendado)', expanded=not st.session_state.cert_cfg.get('pfx_path')):
        _certs_win = _listar_certs_windows()
        if not _certs_win:
            st.warning('Nenhum certificado com chave privada encontrado no store Windows (CurrentUser\\My).')
        else:
            _opcoes = [
                f"{c.get('Subject','').split(',')[0].strip()} — válido até {c.get('NotAfter','')[:10]}"
                for c in _certs_win
            ]
            _idx_sel = st.selectbox('Certificado instalado', range(len(_opcoes)),
                                    format_func=lambda i: _opcoes[i], key='sel_cert_win')
            _cert_sel = _certs_win[_idx_sel]
            st.caption(f'Thumbprint: `{_cert_sel.get("Thumbprint","")}`')
            _col_exp1, _col_exp2 = st.columns([2, 2])
            with _col_exp1:
                _senha_exp = st.text_input(
                    'Senha para o .pfx exportado',
                    type='password',
                    key='inp_senha_export',
                    help='Esta senha será usada no arquivo .pfx gerado e salva nas configurações.',
                )
            with _col_exp2:
                st.write('')
                st.write('')
                if st.button('📤 Exportar e configurar', key='btn_exportar_cert_win', type='primary'):
                    if not _senha_exp:
                        st.error('Informe uma senha para o arquivo .pfx.')
                    else:
                        _thumb = _cert_sel.get('Thumbprint', '')
                        _subj  = _cert_sel.get('Subject', '')
                        _nome_arq = f'cert_{_thumb[:8].lower()}.pfx'
                        _dest_pfx = os.path.join(BASE_DIR, _nome_arq)
                        with st.spinner('Exportando certificado...'):
                            _ok, _err = _exportar_cert_windows(_thumb, _senha_exp, _dest_pfx)
                        if _ok:
                            # Tenta extrair CNPJ do Subject: "CNPJ:12345678000100"
                            import re as _re
                            _m = _re.search(r'CNPJ[:\s]*(\d{14})', _subj)
                            _cnpj_exp = _m.group(1) if _m else ''
                            # Busca CNPJ e UF do ERP se não achou no cert
                            if not _cnpj_exp:
                                _cnpj_exp, _uf_exp = _buscar_dados_loja(st.session_state.db_cfg)
                            else:
                                _, _uf_exp = _buscar_dados_loja(st.session_state.db_cfg)
                            _c = dict(st.session_state.cert_cfg)
                            _c['pfx_path'] = _dest_pfx
                            _c['senha']    = _senha_exp
                            if _cnpj_exp:
                                _c['cnpj'] = _cnpj_exp
                            if _uf_exp:
                                if str(_uf_exp).isdigit():
                                    _c['uf'] = int(_uf_exp)
                                elif _uf_exp in UF_CODIGOS:
                                    _c['uf'] = UF_CODIGOS[_uf_exp]
                            _save_json(CERT_CONFIG, _c)
                            st.session_state.cert_cfg = _c
                            st.success(f'✔ Certificado exportado e configurado: {_nome_arq}')
                            st.rerun()
                        else:
                            st.error(f'Falha ao exportar: {_err}')

    ccfg = st.session_state.cert_cfg
    cs1, cs2, cs3 = st.columns([3, 2, 2])
    with cs1:
        s_pfx = st.text_input('Certificado (.pfx)', value=ccfg.get('pfx_path', ''))
    with cs2:
        s_senha = st.text_input('Senha do certificado', value=ccfg.get('senha', ''), type='password')
    with cs3:
        s_cnpj = st.text_input('CNPJ da empresa', value=ccfg.get('cnpj', ''))

    uf_opts  = sorted(UF_CODIGOS.items())
    uf_siglas = [k for k, _ in uf_opts]
    uf_vals   = [v for _, v in uf_opts]
    uf_atual  = ccfg.get('uf', 32)
    uf_idx    = uf_vals.index(uf_atual) if uf_atual in uf_vals else 0
    s_uf      = st.selectbox('UF da empresa', uf_siglas, index=uf_idx)
    s_uf_cod  = UF_CODIGOS[s_uf]

    if st.button('💾 Salvar configurações do certificado', key='btn_salvar_cert'):
        novo_c = {'pfx_path': s_pfx, 'senha': s_senha,
                  'cnpj': _limpar_cnpj(s_cnpj), 'uf': s_uf_cod,
                  'pasta_nfce': ccfg.get('pasta_nfce', '')}
        _save_json(CERT_CONFIG, novo_c)
        st.session_state.cert_cfg = novo_c
        st.success('Salvo!')

    with st.expander('📂 Pasta de XMLs NFC-e / SAT (opcional — os cupons são lidos do ERP)'):
        s_pasta_nfce = st.text_input(
            'Caminho da pasta com XMLs de NFC-e e SAT',
            value=ccfg.get('pasta_nfce', ''),
            help='Usado apenas como fallback. Os cupons são carregados automaticamente do banco ERP.',
            key='input_pasta_nfce',
        )
        if st.button('💾 Salvar pasta NFC-e/SAT', key='btn_salvar_pasta_nfce'):
            novo_c2 = dict(ccfg)
            novo_c2['pasta_nfce'] = s_pasta_nfce
            _save_json(CERT_CONFIG, novo_c2)
            st.session_state.cert_cfg = novo_c2
            st.success('Pasta salva!')
        if s_pasta_nfce:
            if os.path.isdir(s_pasta_nfce):
                qtd = len([f for f in os.listdir(s_pasta_nfce) if f.lower().endswith('.xml')])
                st.info(f'📂 {qtd} arquivo(s) XML encontrado(s) na pasta.')
            else:
                st.warning('Pasta não encontrada.')

    st.markdown('---')

    cnpj_d = _limpar_cnpj(s_cnpj)
    info_c = get_info_cache(cnpj_d) if cnpj_d else {}

    m1, m2, m3, m4 = st.columns(4)
    nsu_v = info_c.get('ultimo_nsu', '000000000000000')
    m1.metric('Último NSU', nsu_v if nsu_v != '000000000000000' else '(início)')
    m2.metric('NF-e Entrada em cache', len(st.session_state.nfs_xml_mem))
    m3.metric('NF-e Saída em cache', len(st.session_state.nfs_xml_saida_mem))
    m4.metric('Última sincronização', info_c.get('atualizado', '—'))

    # ── Info Task Scheduler ──────────────────────────────────────────────────
    try:
        import subprocess as _sp
        _task_nome = (_loja_ativa.get('task_name', 'MonitorNFe_Framodas')
                      if _loja_ativa else 'MonitorNFe_Framodas')
        _r = _sp.run(
            ['schtasks', '/Query', '/TN', _task_nome, '/FO', 'LIST', '/V'],
            capture_output=True, timeout=5
        )
        _sched_out = _r.stdout.decode('cp850', errors='ignore')
        _prox = '—'
        _ult  = '—'
        for _linha in _sched_out.splitlines():
            _ll = _linha.lower()
            # "Hora da próxima execução" / "Next Run Time"
            if 'xima execu' in _ll or 'next run time' in _ll:
                _prox = _linha.split(':', 1)[-1].strip()
            # "Horário da última execução" / "Last Run Time"
            elif 'ltima execu' in _ll or 'last run time' in _ll:
                _ult = _linha.split(':', 1)[-1].strip()
        st.caption(
            f'🕒 **Monitor automático (Task Scheduler)** — '
            f'Última execução: **{_ult}** | Próxima execução agendada: **{_prox}**'
        )
        # Aviso se última execução foi há mais de 2 horas
        _alerta_task = False
        if _ult and _ult != '—':
            try:
                from datetime import datetime as _dt2
                import re as _re
                # Tenta parsear data/hora no formato dd/mm/yyyy HH:MM:SS ou MM/dd/yyyy HH:MM:SS
                _m = _re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})\s+(\d{1,2}):(\d{2}):(\d{2})', _ult)
                if _m:
                    _g = [int(x) for x in _m.groups()]
                    # Tenta dd/mm/yyyy (padrão PT-BR)
                    try:
                        _dt_ult = _dt2(_g[2], _g[1], _g[0], _g[3], _g[4], _g[5])
                    except ValueError:
                        _dt_ult = _dt2(_g[2], _g[0], _g[1], _g[3], _g[4], _g[5])
                    _diff_h = (_dt2.now() - _dt_ult).total_seconds() / 3600
                    if _diff_h > 2:
                        _alerta_task = True
            except Exception:
                pass
        else:
            _alerta_task = True  # não conseguiu ler a data
        if _alerta_task:
            st.warning(
                f'⚠️ **Monitor automático parado ou inativo há mais de 2 horas!** '
                f'Verifique o Task Scheduler — novas NFs podem não estar sendo capturadas.',
                icon='⚠️'
            )
    except Exception:
        pass

    # ── Ciência da Operação ──────────────────────────────────────────────────
    if cnpj_d:
        _docs_cache = carregar_cache(cnpj_d)
        _res_nfe_ch = extrair_res_nfe_ch(_docs_cache)
        # extrai chaves reais de todos os procNFe no cache (parse do XML se necessário)
        _chaves_com_proc = set()
        for _pd in _docs_cache:
            if 'procNFe' not in _pd.get('schema', '') and 'nfeProc' not in _pd.get('schema', ''):
                continue
            _pch = _pd.get('chave', '')  # disponível para docs via NFeConsulta4
            if _pch and len(_pch) == 44:
                _chaves_com_proc.add(_pch)
                continue
            try:  # para docs DistDFeInt (sem campo chave), extrai do XML
                _pr = ET.fromstring(_pd['xml'])
                _piNFe = _pr.find(f'.//{NS_NFE}infNFe') or _pr.find('.//infNFe')
                if _piNFe is not None:
                    _pc = _piNFe.get('Id', '').replace('NFe', '')
                    if len(_pc) == 44:
                        _chaves_com_proc.add(_pc)
            except Exception:
                pass
        # NF-e ainda genuinamente em formato resumo (sem procNFe no cache)
        _res_sem_proc = [ch for ch in _res_nfe_ch if ch not in _chaves_com_proc]
        # Verifica quais têm ciência expirada (596)
        _exp_path_ci = os.path.join(BASE_DIR, f'{cnpj_d}_ciencia_expirada.json')
        try:
            with open(_exp_path_ci, 'r', encoding='utf-8') as _fe_ci:
                _chaves_exp_ci = set(json.load(_fe_ci))
        except (FileNotFoundError, json.JSONDecodeError):
            _chaves_exp_ci = set()
        _res_expirados = [ch for ch in _res_sem_proc if ch in _chaves_exp_ci]
        _res_baixaveis = [ch for ch in _res_sem_proc if ch not in _chaves_exp_ci]
        # Conta procNFe onde Framodas é a destinatária (NF-e de entrada reais)
        _n_ent_xml = len(filtrar_nfes(_docs_cache, cnpj_d))
        if _res_baixaveis:
            st.info(
                f'📋 **{len(_res_baixaveis)} NF-e ainda como resumo** — o monitor automático '
                f'baixará o XML completo na próxima execução (a cada 65 min). '
                f'({_n_ent_xml} NF-e de entrada com XML completo já em cache)'
            )
        if _res_expirados:
            st.warning(
                f'⚠️ **{len(_res_expirados)} NF-e com ciência expirada** — a SEFAZ não '
                f'entrega mais o XML completo dessas notas (cStat 596). '
                f'Use a **Agência Virtual SEFAZ-ES** abaixo para baixá-las manualmente.'
            )

    # ── Complementar com chaves da Agência Virtual ───────────────────────────
    with st.expander('🔑 Complementar XMLs via Agência Virtual SEFAZ-ES', expanded=False):
        st.info(
            '⚠️ **Por que isso existe?** O serviço DistDFeInt (Receita Federal) distribui '
            'documentos de forma assíncrona e **nem sempre inclui todas as NF-e** do '
            'destinatário — alguns fornecedores emitem pelo sistema estadual e as notas '
            'aparecem na Agência Virtual mas chegam com atraso (ou não chegam) pelo '
            'DistDFeInt. Use esta seção para complementar o cache com as notas faltantes.',
            icon='ℹ️',
        )
        st.markdown('#### Opção 1 — Colar texto (mais rápido)')
        st.markdown(
            '1. Acesse a [Agência Virtual SEFAZ-ES](https://s1-internet.sefaz.es.gov.br) → '
            '**Documentos Fiscais → NF-e → Destinatário → Consultar**\n'
            '2. Na tabela de resultados, selecione tudo (`Ctrl+A`) e copie (`Ctrl+C`)\n'
            '3. Cole aqui abaixo — o app extrai automaticamente todas as chaves de 44 dígitos'
        )
        _txt_cole = st.text_area(
            'Cole o conteúdo copiado da tabela (ou apenas as chaves, uma por linha):',
            height=120,
            key='ta_chaves_cole',
            placeholder='32260341108397000117550010000044421100040974\n32260360798367000124550010000008421167522046\n...',
        )

        st.markdown('#### Opção 2 — Upload de Excel')
        _f_excel = st.file_uploader(
            'Excel baixado da Agência Virtual (.xlsx)',
            type=['xlsx'],
            key='fu_sefaz_excel',
            help='Clique em "Baixar página atual" na Agência Virtual e faça upload aqui.',
        )

        # Extrai chaves de ambas as fontes
        import re as _re_ch
        _chaves_input: list[str] = []
        # Fonte 1: texto colado
        if _txt_cole and _txt_cole.strip():
            _chaves_input += _re_ch.findall(r'\d{44}', _txt_cole)
        # Fonte 2: Excel
        if _f_excel is not None:
            try:
                _df_sefaz = pd.read_excel(_f_excel, dtype=str)
                _df_sefaz.columns = [str(c).strip().upper() for c in _df_sefaz.columns]
                _col_ch = next(
                    (c for c in _df_sefaz.columns if 'CHAVE' in c and 'ACESSO' in c),
                    next((c for c in _df_sefaz.columns if 'CHAVE' in c), None)
                )
                if _col_ch:
                    for _v in _df_sefaz[_col_ch].dropna():
                        _cv = str(_v).strip().replace(' ', '').replace('-', '')
                        if len(_cv) == 44 and _cv.isdigit():
                            _chaves_input.append(_cv)
                else:
                    st.warning(f'Coluna "CHAVE ACESSO" não encontrada no Excel. Colunas: {list(_df_sefaz.columns)}')
            except Exception as _e_xl:
                st.error(f'Erro ao ler Excel: {_e_xl}')
        # Deduplica mantendo ordem
        _chaves_input = list(dict.fromkeys(_chaves_input))

        if _chaves_input:
            # Compara com cache
            _cache_now = carregar_cache(cnpj_d) if cnpj_d else []
            _chaves_cache_now = set()
            for _cpd in _cache_now:
                if 'procNFe' not in _cpd.get('schema','') and 'nfeProc' not in _cpd.get('schema',''):
                    continue
                _cpc = _cpd.get('chave','')
                if _cpc and len(_cpc) == 44:
                    _chaves_cache_now.add(_cpc)
                    continue
                try:
                    _cpr = ET.fromstring(_cpd['xml'])
                    _cpiNFe = _cpr.find(f'.//{NS_NFE}infNFe') or _cpr.find('.//infNFe')
                    if _cpiNFe is not None:
                        _cpc2 = _cpiNFe.get('Id','').replace('NFe','')
                        if len(_cpc2) == 44:
                            _chaves_cache_now.add(_cpc2)
                except Exception:
                    pass
            _faltando = [ch for ch in _chaves_input if ch not in _chaves_cache_now]
            _ja_tem   = [ch for ch in _chaves_input if ch in _chaves_cache_now]
            st.success(
                f'✔ **{len(_chaves_input)} chaves** encontradas | '
                f'**{len(_faltando)} faltando** no cache | '
                f'{len(_ja_tem)} já baixadas'
            )
            if _faltando:
                _df_prev = pd.DataFrame({
                    'Chave': _chaves_input,
                    'Status': ['✔ Já em cache' if ch in _chaves_cache_now else '⬇ Faltando'
                               for ch in _chaves_input],
                })
                st.dataframe(_df_prev, use_container_width=True, hide_index=True, height=200)
                if st.button(
                    f'⬇️ Baixar {len(_faltando)} XML(s) faltando via SEFAZ (DistDFeInt)',
                    key='btn_baixar_excel_chaves', type='primary',
                ):
                    if not s_pfx or not os.path.exists(s_pfx):
                        st.error(f'Certificado não encontrado: {s_pfx}')
                    elif not s_senha:
                        st.warning('Informe a senha do certificado.')
                    elif len(cnpj_d) != 14:
                        st.warning('CNPJ inválido.')
                    else:
                        _logs_exc = []
                        _ph_exc   = st.empty()
                        def _log_exc(msg):
                            _logs_exc.append(msg)
                            _ph_exc.info(msg)
                        _prog_exc = st.progress(0)
                        with st.spinner(f'Baixando {len(_faltando)} XML(s)...'):
                            try:
                                _novos_exc, _erros_exc = baixar_proc_nfe_por_chave(
                                    pfx_path=s_pfx, senha=s_senha,
                                    cnpj=cnpj_d, ch_nfes=_faltando,
                                    log=_log_exc,
                                )
                                _ph_exc.empty()
                                _prog_exc.progress(1.0)
                                if _novos_exc:
                                    _todos_upd = carregar_cache(cnpj_d)
                                    _xmls_upd, _ = ler_xmls_receita(filtrar_nfes(_todos_upd, cnpj_d))
                                    _chaves_proc_upd = {x.get('CHAVE','') for x in _xmls_upd}
                                    for _dcu in _todos_upd:
                                        if 'resNFe' in _dcu.get('schema',''):
                                            _rdu = _res_nfe_to_xml_dict(_dcu['xml'])
                                            if _rdu and _rdu.get('CHAVE','') not in _chaves_proc_upd:
                                                _xmls_upd.append(_rdu)
                                    st.session_state.nfs_xml_mem = _xmls_upd
                                    st.success(
                                        f'✔ {_novos_exc} XML(s) baixado(s)! '
                                        f'Cache atualizado: {len(st.session_state.nfs_xml_mem)} NF-e de entrada.'
                                    )
                                _rate_limited_exc = [ch for ch, msg in _erros_exc if 'RATE_LIMIT' in msg]
                                if _rate_limited_exc:
                                    st.warning(
                                        f'⚠️ Limite SEFAZ atingido (20 consultas/hora). '
                                        f'{_novos_exc} baixado(s), {len(_rate_limited_exc)} restante(s). '
                                        f'Clique no botão novamente em ~1 hora.'
                                    )
                                _outros_erros_exc = [(ch, msg) for ch, msg in _erros_exc if 'RATE_LIMIT' not in msg]
                                if _outros_erros_exc:
                                    with st.expander(f'⚠️ {len(_outros_erros_exc)} erro(s)'):
                                        for _ch_e, _msg_e in _outros_erros_exc:
                                            st.text(f'...{_ch_e[-12:]}: {_msg_e}')
                                if _logs_exc:
                                    with st.expander('📄 Log download'):
                                        for _m in _logs_exc:
                                            st.text(_m)
                            except Exception as _exc_e:
                                _ph_exc.empty()
                                st.error(f'Erro: {_exc_e}')
            else:
                st.success('Todos os XMLs já estão em cache! ✔')

        st.divider()
        st.markdown('#### Opção 3 — Importar XMLs de uma pasta')
        st.markdown(
            '1. Baixe os XMLs desejados na [Agência Virtual SEFAZ-ES](https://s1-internet.sefaz.es.gov.br) '
            '→ **Documentos Fiscais → NF-e → Destinatário → Download XML**\n'
            '2. Faça upload de todos os arquivos `.xml` abaixo'
        )
        _xmls_upload = st.file_uploader(
            'Selecione os arquivos XML',
            type=['xml'],
            accept_multiple_files=True,
            key='fu_xmls_pasta',
        )
        if _xmls_upload:
            if st.button(
                f'📥 Importar {len(_xmls_upload)} XML(s) para o cache',
                key='btn_importar_xmls', type='primary',
            ):
                _importados = 0
                _erros_imp = []
                _novos_imp = []
                for _arq in _xmls_upload:
                    try:
                        _xml_bytes = _arq.read()
                        _xml_str = _xml_bytes.decode('utf-8')
                        _root_imp = ET.fromstring(_xml_str)
                        _inf_imp = (_root_imp.find(f'.//{NS_NFE}infNFe')
                                    or _root_imp.find('.//infNFe'))
                        if _inf_imp is None:
                            _erros_imp.append((_arq.name, 'infNFe não encontrado no XML'))
                            continue
                        _ch_imp = _inf_imp.get('Id', '').replace('NFe', '')
                        if len(_ch_imp) != 44:
                            _erros_imp.append((_arq.name, f'Chave inválida: {_ch_imp}'))
                            continue
                        _novos_imp.append({
                            'nsu': f'IMP_{_ch_imp}',
                            'schema': 'procNFe_v4.00.xsd',
                            'xml': _xml_str,
                        })
                        _importados += 1
                    except Exception as _e_imp:
                        _erros_imp.append((_arq.name, str(_e_imp)))
                if _novos_imp and cnpj_d:
                    salvar_cache(cnpj_d, _novos_imp)
                    _todos_upd3 = carregar_cache(cnpj_d)
                    _xmls_upd3, _ = ler_xmls_receita(filtrar_nfes(_todos_upd3, cnpj_d))
                    _chaves_proc_upd3 = {x.get('CHAVE', '') for x in _xmls_upd3}
                    for _dcu3 in _todos_upd3:
                        if 'resNFe' in _dcu3.get('schema', ''):
                            _rdu3 = _res_nfe_to_xml_dict(_dcu3['xml'])
                            if _rdu3 and _rdu3.get('CHAVE', '') not in _chaves_proc_upd3:
                                _xmls_upd3.append(_rdu3)
                    st.session_state.nfs_xml_mem = _xmls_upd3
                    st.success(
                        f'✔ {_importados} XML(s) importado(s)! '
                        f'Cache atualizado: {len(st.session_state.nfs_xml_mem)} NF-e de entrada.'
                    )
                if _erros_imp:
                    with st.expander(f'⚠️ {len(_erros_imp)} erro(s) na importação'):
                        for _nm_e, _msg_e in _erros_imp:
                            st.text(f'{_nm_e}: {_msg_e}')

    st.markdown('---')
    cb1, cb2 = st.columns([3, 1])

    with cb1:
        if st.button('🔄 Buscar Receita Federal', type='primary',
                     use_container_width=True, key='btn_buscar_rf'):
            if not s_pfx or not os.path.exists(s_pfx):
                st.error(f'Certificado não encontrado: {s_pfx}')
            elif not s_senha:
                st.warning('Informe a senha do certificado.')
            elif len(cnpj_d) != 14:
                st.warning('CNPJ inválido (deve ter 14 dígitos).')
            else:
                logs_b = []
                log_ph = st.empty()

                def log_fn(msg):
                    logs_b.append(msg)
                    log_ph.info(msg)

                with st.spinner('Consultando Receita Federal...'):
                    try:
                        docs_raw, ult_nsu = consultar_distdfe(
                            pfx_path=s_pfx, senha_pfx=s_senha,
                            cnpj=cnpj_d, uf_autor=s_uf_cod, log=log_fn,
                        )
                        nfes_xml = filtrar_nfes(docs_raw, cnpj_d)
                        xmls, erros = ler_xmls_receita(nfes_xml)

                        # Após download incremental, carrega o cache completo
                        todos_cache = carregar_cache(cnpj_d)

                        # Baixa automaticamente o XML completo de qualquer resNFe
                        # que ainda não tenha procNFe no cache (sem precisar de Ciência)
                        _res_chaves = extrair_res_nfe_ch(todos_cache)
                        _chaves_proc_cache = set()
                        for _dp in todos_cache:
                            if 'procNFe' not in _dp.get('schema', '') and 'nfeProc' not in _dp.get('schema', ''):
                                continue
                            _pch = _dp.get('chave', '')
                            if _pch and len(_pch) == 44:
                                _chaves_proc_cache.add(_pch)
                                continue
                            try:
                                _pr = ET.fromstring(_dp['xml'])
                                _piNFe = _pr.find(f'.//{NS_NFE}infNFe') or _pr.find('.//infNFe')
                                if _piNFe is not None:
                                    _pc = _piNFe.get('Id', '').replace('NFe', '')
                                    if len(_pc) == 44:
                                        _chaves_proc_cache.add(_pc)
                            except Exception:
                                pass
                        # Exclui chaves com ciência expirada (596)
                        _exp_path = os.path.join(BASE_DIR, f'{cnpj_d}_ciencia_expirada.json')
                        try:
                            with open(_exp_path, 'r', encoding='utf-8') as _fe:
                                _chaves_exp = set(json.load(_fe))
                        except (FileNotFoundError, json.JSONDecodeError):
                            _chaves_exp = set()
                        _chaves_para_baixar = [ch for ch in _res_chaves
                                               if ch not in _chaves_proc_cache and ch not in _chaves_exp]
                        if _chaves_para_baixar:
                            log_fn(f'Baixando {len(_chaves_para_baixar)} XML(s) completo(s) por chave...')
                            _nov_ch, _err_ch = baixar_proc_nfe_por_chave(
                                pfx_path=s_pfx, senha=s_senha,
                                cnpj=cnpj_d, ch_nfes=_chaves_para_baixar, log=log_fn,
                            )
                            if _nov_ch:
                                log_fn(f'{_nov_ch} XML(s) completo(s) baixados com sucesso.')
                                todos_cache = carregar_cache(cnpj_d)  # recarrega com os novos
                            if _err_ch:
                                log_fn(f'Aviso: {len(_err_ch)} chave(s) não puderam ser baixadas.')

                        xmls_total, _ = ler_xmls_receita(filtrar_nfes(todos_cache, cnpj_d))
                        # Inclui resNFe (resumos) que ainda não viraram procNFe (fallback)
                        _chaves_proc = {x.get('CHAVE', '') for x in xmls_total}
                        for _dc in todos_cache:
                            if 'resNFe' in _dc.get('schema', ''):
                                _rd = _res_nfe_to_xml_dict(_dc['xml'])
                                if _rd and _rd.get('CHAVE', '') not in _chaves_proc:
                                    xmls_total.append(_rd)
                        st.session_state.nfs_xml_mem = xmls_total if xmls_total else xmls

                        xmls_saida, _ = ler_xmls_receita(filtrar_nfes_saida(todos_cache, cnpj_d))
                        st.session_state.nfs_xml_saida_mem = xmls_saida

                        st.session_state.sefaz_logs = logs_b
                        log_ph.empty()
                        novos = len(xmls)
                        total_ent = len(st.session_state.nfs_xml_mem)
                        total_sai = len(st.session_state.nfs_xml_saida_mem)
                        st.success(
                            f'✔ {len(docs_raw)} doc(s) novos baixados | '
                            f'{novos} NF-e novas | '
                            f'**{total_ent} NF-e entrada** | '
                            f'**{total_sai} NF-e saída** em cache | NSU: {ult_nsu}'
                        )
                        if erros:
                            st.warning(f'{len(erros)} XML(s) com erro de leitura.')
                        if docs_raw == [] or novos == 0:
                            st.info('Nenhum documento novo desde o último NSU. Usando cache local completo.')
                    except _ConsumoIndevido as e:
                        st.session_state.sefaz_logs = logs_b
                        log_ph.empty()
                        cache_docs = e.cache
                        if cache_docs:
                            # Mesmo com rate limit no distNSU, tenta baixar XMLs por chave
                            # (consChNFe usa modo diferente e pode não ter o mesmo bloqueio)
                            _res_ch_rl = extrair_res_nfe_ch(cache_docs)
                            _proc_ch_rl = set()
                            for _drl in cache_docs:
                                if 'procNFe' not in _drl.get('schema', '') and 'nfeProc' not in _drl.get('schema', ''):
                                    continue
                                _pch_rl = _drl.get('chave', '')
                                if _pch_rl and len(_pch_rl) == 44:
                                    _proc_ch_rl.add(_pch_rl)
                                    continue
                                try:
                                    _pr_rl = ET.fromstring(_drl['xml'])
                                    _pi_rl = _pr_rl.find(f'.//{NS_NFE}infNFe') or _pr_rl.find('.//infNFe')
                                    if _pi_rl is not None:
                                        _pc_rl = _pi_rl.get('Id', '').replace('NFe', '')
                                        if len(_pc_rl) == 44:
                                            _proc_ch_rl.add(_pc_rl)
                                except Exception:
                                    pass
                            _baixar_rl = [ch for ch in _res_ch_rl if ch not in _proc_ch_rl]
                            # Exclui chaves com ciência expirada (596)
                            _exp_path_rl = os.path.join(BASE_DIR, f'{cnpj_d}_ciencia_expirada.json')
                            try:
                                with open(_exp_path_rl, 'r', encoding='utf-8') as _fe_rl:
                                    _chaves_exp_rl = set(json.load(_fe_rl))
                            except (FileNotFoundError, json.JSONDecodeError):
                                _chaves_exp_rl = set()
                            _expirados_rl = [ch for ch in _baixar_rl if ch in _chaves_exp_rl]
                            _baixar_rl = [ch for ch in _baixar_rl if ch not in _chaves_exp_rl]
                            _baixados_rl = 0
                            if _baixar_rl:
                                with st.spinner(f'DistDFeInt bloqueado — tentando baixar {len(_baixar_rl)} XML(s) por chave...'):
                                    try:
                                        # Envia Ciência da Operação primeiro para garantir acesso
                                        manifestar_ciencia(
                                            pfx_path=s_pfx, senha=s_senha, cnpj=cnpj_d,
                                            uf_autor=s_uf_cod, ch_nfes=_baixar_rl, log=log_fn,
                                        )
                                    except Exception:
                                        pass
                                    try:
                                        _nov_rl, _err_rl = baixar_proc_nfe_por_chave(
                                            pfx_path=s_pfx, senha=s_senha,
                                            cnpj=cnpj_d, ch_nfes=_baixar_rl, log=log_fn,
                                        )
                                        _baixados_rl = _nov_rl
                                        if _nov_rl:
                                            cache_docs = carregar_cache(cnpj_d)  # recarrega com novos
                                    except Exception:
                                        pass

                            xmls_c, _ = ler_xmls_receita(filtrar_nfes(cache_docs, cnpj_d))
                            _chaves_proc_c = {x.get('CHAVE', '') for x in xmls_c}
                            for _dc in cache_docs:
                                if 'resNFe' in _dc.get('schema', ''):
                                    _rd = _res_nfe_to_xml_dict(_dc['xml'])
                                    if _rd and _rd.get('CHAVE', '') not in _chaves_proc_c:
                                        xmls_c.append(_rd)
                            st.session_state.nfs_xml_mem = xmls_c
                            xmls_cs, _ = ler_xmls_receita(filtrar_nfes_saida(cache_docs, cnpj_d))
                            st.session_state.nfs_xml_saida_mem = xmls_cs
                            _info_ci = get_info_cache(cnpj_d)
                            _msg_rl = (
                                f'A SEFAZ bloqueou a consulta (rate limit por CNPJ). '
                                f'Usando cache local: **{len(cache_docs)} doc(s)** | '
                                f'Última sincronização: {_info_ci.get("atualizado", "—")}'
                            )
                            if _baixados_rl:
                                _msg_rl += f' | ✔ **{_baixados_rl} XML(s) completo(s) baixados por chave**'
                            elif _baixar_rl:
                                _msg_rl += f' | ⏳ {len(_baixar_rl)} resumo(s) ainda aguardando — consChNFe também bloqueado'
                            if _expirados_rl:
                                _msg_rl += f' | ⚠️ {len(_expirados_rl)} resumo(s) com ciência expirada (irrecuperáveis via SEFAZ)'
                            st.info(_msg_rl, icon='ℹ️')
                        else:
                            st.warning('SEFAZ bloqueou a consulta (rate limit) e não há cache local disponível. Tente mais tarde.', icon='⏳')
                    except Exception as e:
                        st.session_state.sefaz_logs = logs_b
                        log_ph.empty()
                        st.error(f'Erro SEFAZ: {e}')

    with cb2:
        if st.button('♻️ Resetar NSU', use_container_width=True, key='btn_reset_nsu',
                     help='Zera o NSU para rebaixar documentos dos últimos 90 dias. O cache de XMLs é preservado.'):
            if not st.session_state.get('_reset_confirm'):
                st.session_state['_reset_confirm'] = True
                st.warning('⚠️ O NSU será zerado e a SEFAZ enviará todos os documentos '
                           'dos últimos 90 dias novamente. Os XMLs já em cache serão preservados. '
                           'Clique em **♻️ Resetar NSU** novamente para confirmar.')
            else:
                st.session_state.pop('_reset_confirm', None)
                if cnpj_d:
                    salvar_ultimo_nsu(cnpj_d, '000000000000000')
                    st.session_state.nfs_xml_mem = []
                    st.session_state.nfs_xml_saida_mem = []
                    st.success('NSU zerado. XMLs em cache preservados. Clique em Buscar para baixar tudo novamente.')
                else:
                    st.warning('Informe o CNPJ.')

    if st.session_state.sefaz_logs:
        with st.expander('📄 Log da última consulta'):
            for msg in st.session_state.sefaz_logs:
                st.text(msg)

# ══════════════════════════════════════════════════════════════════════════════
# TAB: CONFERÊNCIA
# ══════════════════════════════════════════════════════════════════════════════
with tab_conf:
    # ── Controles ─────────────────────────────────────────────────────────────
    cc1, cc2, cc3, cc4 = st.columns([1.3, 1.3, 2.0, 1.4])
    with cc1:
        data_ini = st.date_input('De', value=st.session_state.data_ini,
                                 format='DD/MM/YYYY', key='w_ini')
        st.session_state.data_ini = data_ini
    with cc2:
        data_fim = st.date_input('Até', value=st.session_state.data_fim,
                                 format='DD/MM/YYYY', key='w_fim')
        st.session_state.data_fim = data_fim
    with cc3:
        tipo_data_erp = st.radio(
            'Data de referência no ERP',
            ['Emissão da NF', 'Entrada no ERP'],
            horizontal=True, key='r_tipo_data',
            help='Emissão = data da NF-e no XML. Entrada = data do lançamento no ERP.',
        )
        _usar_emissao_erp = (tipo_data_erp == 'Emissão da NF')
    with cc4:
        st.write('')
        st.write('')
        conferir_btn = st.button('▶ Conferir', type='primary',
                                 use_container_width=True, key='btn_conferir')

    # ── Aviso se não há XMLs ───────────────────────────────────────────────────
    if not st.session_state.nfs_xml_mem and not st.session_state.resultados:
        st.info(
            '**Sem XMLs da Receita Federal.** '
            'Vá até a aba 🏛️ **Receita Federal** e clique em 🔄 Buscar para baixar as NF-e.',
            icon='📥',
        )

    if conferir_btn:
        nfs_xml = st.session_state.nfs_xml_mem
        if not nfs_xml:
            st.error('Nenhum XML da Receita Federal. Acesse a aba Receita Federal e clique em Buscar.')
            st.stop()

        _d0, _d1 = str(data_ini), str(data_fim)
        nfs_xml_periodo = [x for x in nfs_xml if _d0 <= (x.get('EMISSAO') or '')[:10] <= _d1]
        if not nfs_xml_periodo:
            st.warning(
                f'Sem XMLs com emissão entre {data_ini.strftime("%d/%m/%Y")} '
                f'e {data_fim.strftime("%d/%m/%Y")}. Tente um período mais amplo.',
                icon='⚠️',
            )
            st.stop()

        with st.spinner('Consultando ERP...'):
            try:
                con = conectar_erp_loja(_loja_ativa, st.session_state.db_cfg)
                if _loja_ativa and _loja_ativa.get('tabela_entrada') == 'NFISCAL':
                    nfs_erp = buscar_nfs_nfiscal_entrada(
                        con, data_ini, data_fim,
                        loja_id=_loja_ativa['id'],
                        usar_emissao=_usar_emissao_erp,
                    )
                else:
                    nfs_erp = buscar_nfs_erp(con, data_ini, data_fim,
                                             usar_emissao=_usar_emissao_erp,
                                             loja_id=_loja_ativa['id'] if _loja_ativa else None)
                con.close()
            except Exception as e:
                st.error(f'Erro ao conectar ao ERP:\n\n{e}')
                st.stop()

        st.session_state.resultados = conferir(nfs_erp, nfs_xml_periodo)

    # ── Exibição dos resultados ──────────────────────────────────────────────
    if st.session_state.resultados:
        res = st.session_state.resultados

        # Separa: XMLs da Receita × entradas só no ERP
        res_xml = [r for r in res if r['Status'] != STATUS_SO_ERP]
        res_erp = [r for r in res if r['Status'] == STATUS_SO_ERP]

        n_total_xml = len(res_xml)
        n_lancadas  = sum(1 for r in res_xml if r['Status'] in (STATUS_OK, STATUS_RESUMO))
        n_nao_lanc  = sum(1 for r in res_xml if r['Status'] == STATUS_SO_XML)
        n_diverg    = sum(1 for r in res_xml if r['Status'] == STATUS_DIVERGENCIA)

        # ── Métricas ──────────────────────────────────────────────────────────
        mc1, mc2, mc3, mc4 = st.columns(4)
        mc1.metric('Total de XMLs SEFAZ', n_total_xml)
        mc2.metric('✅ Lançadas no ERP', n_lancadas)
        mc3.metric('❌ Não lançadas no ERP', n_nao_lanc)
        mc4.metric('⚠️ Divergência de valor', n_diverg)

        n_res = sum(1 for r in res_xml if r['Status'] == STATUS_RESUMO)
        if n_res:
            st.info(
                f'**{n_res} nota(s)** foram conferidas pelo resumo da Receita Federal (resNFe). '
                'O valor total foi comparado — detalhes de IPI, frete e ICMS disponíveis quando o XML completo chegar.',
                icon='ℹ️',
            )

        # ── Filtro ────────────────────────────────────────────────────────────
        _opcoes_filtro = {
            'Todas'        : None,
            'Não lançadas' : [STATUS_SO_XML],
            'Divergências' : [STATUS_DIVERGENCIA],
            'Lançadas'     : [STATUS_OK, STATUS_RESUMO],
        }
        _sel_filtro = st.radio(
            'Exibir', list(_opcoes_filtro.keys()),
            horizontal=True, key='r_filtro_conf',
        )
        _status_filtro = _opcoes_filtro[_sel_filtro]
        res_view = (
            [r for r in res_xml if r['Status'] in _status_filtro]
            if _status_filtro else res_xml
        )

        # ── Tabela ────────────────────────────────────────────────────────────
        _LABEL_STATUS = {
            STATUS_OK         : 'Lançada',
            STATUS_RESUMO     : 'Lançada',
            STATUS_DIVERGENCIA: 'Divergência',
            STATUS_SO_XML     : 'Não lançada',
        }
        _COR_LINHA = {
            STATUS_OK         : 'background-color: #C6EFCE',
            STATUS_RESUMO     : 'background-color: #C6EFCE',
            STATUS_DIVERGENCIA: 'background-color: #FFEB9C',
            STATUS_SO_XML     : 'background-color: #FCE4D6',
        }
        _fmt_brl_c = lambda v: (
            f'R$ {float(v):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
        )
        df_view = pd.DataFrame([{
            'Situação'       : _LABEL_STATUS.get(r['Status'], r['Status']),
            'Nº NF'          : r['Número'],
            'Série'          : r['Série'],
            'Dt. Emissão'    : r['Dt. Emissão'],
            'Dt. Entrada ERP': r['Dt. Entrada'],
            'Fornecedor'     : r['Fornecedor XML'] or r['Fornecedor ERP'],
            'CNPJ'           : r['CNPJ'],
            'Valor NF'       : float(r['Total XML']) if r['Total XML'] else float(r['Total ERP']),
            'Observação'     : r['Divergências'],
        } for r in res_view]).reset_index(drop=True)

        _cores = [_COR_LINHA.get(r['Status'], '') for r in res_view]

        def _colorir_conf(row):
            c = _cores[row.name] if row.name < len(_cores) else ''
            return [c] * len(row)

        st.dataframe(
            df_view.style.apply(_colorir_conf, axis=1).format({'Valor NF': _fmt_brl_c}),
            use_container_width=True,
            height=460,
        )

        # ── Entradas no ERP sem XML (colapsado) ───────────────────────────────
        if res_erp:
            with st.expander(
                f'📁 Entradas no ERP sem XML na Receita Federal ({len(res_erp)})',
                expanded=False,
            ):
                st.caption(
                    'Estas entradas estão lançadas no ERP mas não foram encontradas '
                    'entre os XMLs baixados da SEFAZ no período. '
                    'Pode ser lançamento manual, período diferente ou nota não eletrônica.'
                )
                df_erp_v = pd.DataFrame([{
                    'Nº NF'          : r['Número'],
                    'Série'          : r['Série'],
                    'Dt. Emissão'    : r['Dt. Emissão'],
                    'Dt. Entrada ERP': r['Dt. Entrada'],
                    'Fornecedor'     : r['Fornecedor ERP'],
                    'CNPJ'           : r['CNPJ'],
                    'Valor NF'       : float(r['Total ERP']),
                } for r in res_erp])
                st.dataframe(
                    df_erp_v.style.format({'Valor NF': _fmt_brl_c}),
                    use_container_width=True,
                )

        # ── Exportar Excel ────────────────────────────────────────────────────
        st.markdown('---')
        excel_data = gerar_excel(res)
        nome_arq = (
            f'conferencia_entrada_'
            f'{st.session_state.data_ini.strftime("%d%m%Y")}_'
            f'{st.session_state.data_fim.strftime("%d%m%Y")}.xlsx'
        )
        st.download_button(
            label='📊 Exportar Excel',
            data=excel_data,
            file_name=nome_arq,
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            type='primary',
            key='btn_excel',
        )

# ══════════════════════════════════════════════════════════════════════════════
# TAB: NF DE SAÍDA
# ══════════════════════════════════════════════════════════════════════════════
with tab_saida:
    nfs_saida_xml = st.session_state.nfs_xml_saida_mem
    _fmt_brl_s = lambda v: f'R$ {float(v):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
    _db_local_saida = bool(_loja_ativa and _loja_ativa.get('db_host'))

    # ── Controles ─────────────────────────────────────────────────────────────
    _sc1, _sc2, _sc3 = st.columns([1.3, 1.3, 3])
    with _sc1:
        sl_ini = st.date_input('De', value=st.session_state.data_ini,
                               format='DD/MM/YYYY', key='sl_ini_saida')
    with _sc2:
        sl_fim = st.date_input('Até', value=st.session_state.data_fim,
                               format='DD/MM/YYYY', key='sl_fim_saida')
    with _sc3:
        _sl_busca = st.text_input('Buscar (cliente, CNPJ, número…)',
                                  key='sl_busca_saida', placeholder='Filtrar...')

    if st.button('🔍 Buscar NF-e de Saída', key='btn_buscar_saida_erp_lista', type='primary'):
        _db_cfg_sl = st.session_state.db_cfg
        if not _db_cfg_sl or not _db_cfg_sl.get('host'):
            st.error('Configure a conexão com o ERP na aba Configurações.')
        else:
            with st.spinner('Buscando NF-e de saída no ERP...'):
                try:
                    _con_sl = conectar_erp_loja(_loja_ativa, _db_cfg_sl)
                    _cnpj_sl = _limpar_cnpj(st.session_state.cert_cfg.get('cnpj', ''))
                    _lid_sl = _loja_ativa['id'] if _loja_ativa else None
                    st.session_state.nfs_erp_saida_lista = buscar_nfs_saida_erp(
                        _con_sl, sl_ini, sl_fim,
                        loja_id=None if _db_local_saida else _lid_sl,
                        cnpj_loja=None if _db_local_saida else (_cnpj_sl or None))
                    _con_sl.close()
                except Exception as _ex_sl:
                    st.error(f'Erro ao consultar ERP: {_ex_sl}')

    # ── Resultado ─────────────────────────────────────────────────────────────
    _erp_sl = st.session_state.get('nfs_erp_saida_lista', [])
    if _erp_sl:
        df_sl = pd.DataFrame(_erp_sl)
        if _sl_busca.strip():
            _tm_sl = _sl_busca.strip().lower()
            df_sl = df_sl[df_sl.apply(lambda row: any(_tm_sl in str(v).lower() for v in row), axis=1)]

        _cols_sl = ['NUMERO', 'SERIE', 'DT_EMISSAO', 'CLIENTE', 'CNPJ_CLI',
                    'TOTAL_NF', 'TOTAL_PROD', 'TOTAL_IPI', 'TOTAL_ICMS', 'TOTAL_ST']
        _cols_sl = [c for c in _cols_sl if c in df_sl.columns]
        _rename_sl = {
            'NUMERO': 'Número', 'SERIE': 'Série', 'DT_EMISSAO': 'Emissão',
            'CLIENTE': 'Cliente', 'CNPJ_CLI': 'CNPJ',
            'TOTAL_NF': 'Total NF', 'TOTAL_PROD': 'Produtos',
            'TOTAL_IPI': 'IPI', 'TOTAL_ICMS': 'ICMS', 'TOTAL_ST': 'ST',
        }
        _fmt_sl = {c: _fmt_brl_s for c in
                   ['TOTAL_NF', 'TOTAL_PROD', 'TOTAL_IPI', 'TOTAL_ICMS', 'TOTAL_ST']
                   if c in df_sl.columns}

        _sm1_sl, _sm2_sl = st.columns(2)
        _sm1_sl.metric('Total NF-e', len(df_sl))
        _sm2_sl.metric('Total geral',
                       _fmt_brl_s(sum(float(v) for v in df_sl['TOTAL_NF']))
                       if 'TOTAL_NF' in df_sl.columns and len(df_sl) > 0 else 'R$ 0,00')

        st.dataframe(df_sl[_cols_sl].rename(columns=_rename_sl).style.format(_fmt_sl),
                     use_container_width=True, height=500)

        _rows_exp_sl = [{k: v for k, v in r.items() if k != 'XML'} for r in _erp_sl]
        _sc_dl1, _sc_dl2, _sc_dl3 = st.columns(3)
        with _sc_dl1:
            st.download_button(
                label='📊 Exportar Excel',
                data=gerar_excel(_rows_exp_sl),
                file_name=f'nf_saida_{sl_ini.strftime("%d%m%Y")}_{sl_fim.strftime("%d%m%Y")}.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                type='primary', key='btn_excel_saida_erp',
                use_container_width=True,
            )
        with _sc_dl2:
            _xmls_sai = [r['XML'] for r in _erp_sl if r.get('XML')]
            _sem_xml_sai = [r for r in _erp_sl if not r.get('XML')]
            if _xmls_sai:
                import zipfile as _zf, io as _io2
                _zbuf = _io2.BytesIO()
                with _zf.ZipFile(_zbuf, 'w', _zf.ZIP_DEFLATED) as _zzip:
                    for _i, _xstr in enumerate(_xmls_sai, 1):
                        _nm = (df_sl.iloc[_i-1]['Número'] if _i-1 < len(df_sl) else _i) if 'Número' in (df_sl.columns if isinstance(df_sl, pd.DataFrame) else []) else _i
                        _ch = _erp_sl[_i-1].get('CHAVE', '') or f'nfe_saida_{_i:04d}'
                        _zzip.writestr(f'{_ch or f"nfe_{_i:04d}"}.xml', _xstr)
                st.download_button(
                    f'📦 Baixar XMLs ({len(_xmls_sai)})',
                    data=_zbuf.getvalue(),
                    file_name=f'xmls_saida_{sl_ini.strftime("%d%m%Y")}_{sl_fim.strftime("%d%m%Y")}.zip',
                    mime='application/zip',
                    key='btn_zip_saida',
                    use_container_width=True,
                )
        if _sem_xml_sai:
            _cnpj_loja_sai = _loja_ativa.get('cnpj', '') if _loja_ativa else ''
            _cache_found_sai = buscar_xml_no_cache(_sem_xml_sai, _cnpj_loja_sai) if SEFAZ_OK else {}
            _nomes_sem_sai = []
            for _si, _sr in enumerate(_sem_xml_sai):
                _lbl = f"NF {_sr.get('NUMERO','')} (série {_sr.get('SERIE','')})"
                if _si in _cache_found_sai:
                    _lbl += ' ✅ encontrado no cache SEFAZ'
                else:
                    _lbl += ' ❌ não encontrado no cache'
                _nomes_sem_sai.append(_lbl)
            _n_found_sai = len(_cache_found_sai)
            _n_miss_sai  = len(_sem_xml_sai) - _n_found_sai
            _msg_sai = (
                f'ℹ️ {len(_sem_xml_sai)} NF(s) sem XML no ERP (lançadas manualmente): '
                + '; '.join(_nomes_sem_sai) + '.'
            )
            if _n_miss_sai > 0:
                _msg_sai += f' {_n_miss_sai} não encontrada(s) no cache SEFAZ — use a Agência Virtual para recuperar.'
            st.info(_msg_sai, icon=None)
            if _n_found_sai > 0:
                import zipfile as _zf_s2, io as _io_s2
                _zbuf_s2 = _io_s2.BytesIO()
                with _zf_s2.ZipFile(_zbuf_s2, 'w', _zf_s2.ZIP_DEFLATED) as _zz_s2:
                    for _si, _cf in _cache_found_sai.items():
                        _ch_s2 = _cf['chave'] or f'nfe_saida_sem_{_si+1:04d}'
                        _zz_s2.writestr(f'{_ch_s2}.xml', _cf['xml'])
                st.download_button(
                    f'📦 Baixar XMLs recuperados do cache ({_n_found_sai})',
                    data=_zbuf_s2.getvalue(),
                    file_name=f'xmls_saida_recuperados_{sl_ini.strftime("%d%m%Y")}.zip',
                    mime='application/zip',
                    key='btn_zip_saida_recuperados',
                    use_container_width=True,
                )
        with _sc_dl3:
            if st.button('📄 Gerar DANFEs (.zip)', key='btn_danfe_saida',
                         use_container_width=True):
                from brazilfiscalreport.danfe import Danfe as _Danfe
                _pdf_buf = io.BytesIO()
                _ger_s = 0; _sem_s = 0
                with zipfile.ZipFile(_pdf_buf, 'w', zipfile.ZIP_DEFLATED) as _pzf:
                    for _r in _erp_sl:
                        _xml_s = _r.get('XML', '')
                        _ch_s  = _r.get('CHAVE', '') or f'nfe_{_r.get("NUMERO","")}'
                        if not _xml_s:
                            _sem_s += 1; continue
                        try:
                            _pdf_obj = _Danfe(xml=_xml_s)
                            _pdf_bytes = _pdf_obj.output()
                            _pzf.writestr(f'{_ch_s}.pdf', bytes(_pdf_bytes))
                            _ger_s += 1
                        except Exception as _exc_d:
                            _pzf.writestr(f'erro_{_ch_s}.txt', f'Erro: {_exc_d}')
                if _ger_s == 0:
                    st.warning(f'Nenhum DANFE gerado. {_sem_s} NF-e sem XML no banco.')
                else:
                    st.session_state['_pdf_zip_bytes_saida'] = _pdf_buf.getvalue()
                    st.session_state['_pdf_zip_name_saida']  = (
                        f'danfes_saida_{sl_ini.strftime("%d%m%Y")}_{sl_fim.strftime("%d%m%Y")}.zip'
                    )
                    if _sem_s > 0:
                        st.info(f'{_ger_s} DANFE(s) gerados. {_sem_s} NF-e sem XML foram ignoradas.', icon='ℹ️')
            if st.session_state.get('_pdf_zip_bytes_saida'):
                st.download_button(
                    label='⬇️ Clique para baixar DANFEs',
                    data=st.session_state['_pdf_zip_bytes_saida'],
                    file_name=st.session_state.get('_pdf_zip_name_saida', 'danfes_saida.zip'),
                    mime='application/zip',
                    key='btn_pdf_dl_saida', use_container_width=True,
                )


# ══════════════════════════════════════════════════════════════════════════════
def _classificar_cfop(cfop):
    """Classifica CFOP na perspectiva do destinatario (comprador)."""
    if not cfop or (isinstance(cfop, float)):
        return ''
    c = str(cfop).replace('.', '').strip()
    if len(c) < 4 or not c.isdigit():
        return ''
    prefixo = c[0]
    centena = int(c[1])
    _ESP = {
        '5101':'Compra','5102':'Compra','5103':'Compra','5104':'Compra','5105':'Compra','5106':'Compra',
        '6101':'Compra','6102':'Compra','6103':'Compra','6104':'Compra','6105':'Compra','6106':'Compra','6107':'Compra','6108':'Compra','6109':'Compra',
        '5111':'Compra','5112':'Compra','5113':'Compra','5116':'Compra','5117':'Compra','5118':'Compra','5119':'Compra','5120':'Compra','5122':'Compra',
        '6111':'Compra','6112':'Compra','6113':'Compra','6116':'Compra','6117':'Compra','6118':'Compra','6119':'Compra','6120':'Compra','6122':'Compra',
        '5401':'Compra ST','5402':'Compra ST','5403':'Compra ST','5405':'Compra ST',
        '6401':'Compra ST','6402':'Compra ST','6403':'Compra ST','6404':'Compra ST','6405':'Compra ST',
        '5151':'Transferencia','5152':'Transferencia','5153':'Transferencia','5155':'Transferencia','5156':'Transferencia',
        '6151':'Transferencia','6152':'Transferencia','6153':'Transferencia','6155':'Transferencia','6156':'Transferencia',
        '5408':'Transferencia ST','5409':'Transferencia ST',
        '6408':'Transferencia ST','6409':'Transferencia ST',
        '5201':'Devolucao','5202':'Devolucao','5208':'Devolucao','5209':'Devolucao',
        '6201':'Devolucao','6202':'Devolucao','6208':'Devolucao','6209':'Devolucao',
        '5410':'Devolucao ST','5411':'Devolucao ST',
        '6410':'Devolucao ST','6411':'Devolucao ST',
        '5551':'Imobilizado','5553':'Imobilizado','6551':'Imobilizado','6553':'Imobilizado',
        '5556':'Uso/Consumo','5557':'Uso/Consumo','6556':'Uso/Consumo','6557':'Uso/Consumo',
        '5910':'Bonificacao','6910':'Bonificacao',
        '5911':'Amostra gratis','6911':'Amostra gratis',
        '5933':'Servico','6933':'Servico','5932':'Servico','6932':'Servico',
        '5352':'Energia/Agua','5353':'Energia/Agua',
        '5949':'Outros','6949':'Outros',
        '5902':'Remessa','5903':'Remessa','5904':'Remessa','5905':'Remessa','5906':'Remessa','5907':'Remessa','5908':'Remessa','5909':'Remessa',
        '6902':'Remessa','6903':'Remessa','6904':'Remessa','6905':'Remessa','6906':'Remessa','6907':'Remessa','6908':'Remessa','6909':'Remessa',
        '1102':'Compra','2102':'Compra','1101':'Compra','2101':'Compra',
        '1403':'Compra ST','2403':'Compra ST','1401':'Compra ST','2401':'Compra ST',
        '1152':'Transferencia','2152':'Transferencia',
        '1202':'Devolucao','2202':'Devolucao',
        '1556':'Uso/Consumo','2556':'Uso/Consumo',
        '1551':'Imobilizado','2551':'Imobilizado',
        '1910':'Bonificacao','2910':'Bonificacao',
        '1933':'Servico','2933':'Servico',
        '1949':'Outros','2949':'Outros',
    }
    if c in _ESP:
        return _ESP[c]
    if prefixo in ('5', '6'):
        if centena == 1: return 'Compra'
        if centena == 2: return 'Devolucao'
        if centena == 4: return 'Compra ST'
        if centena == 5: return 'Uso/Consumo'
        if centena == 9: return 'Remessa/Outros'
        return 'Compra'
    if prefixo in ('1', '2'):
        if centena == 1: return 'Compra'
        if centena == 2: return 'Devolucao'
        if centena == 4: return 'Compra ST'
        if centena == 5: return 'Uso/Consumo'
        if centena == 9: return 'Outros'
        return 'Compra'
    return 'Outros'


# TAB: NF DE ENTRADA
# ══════════════════════════════════════════════════════════════════════════════
with tab_entrada:
    st.subheader('NF-e de Entrada — XMLs da Receita Federal (SEFAZ)')
    nfs_ent_xml = st.session_state.nfs_xml_mem
    if not nfs_ent_xml:
        st.info(
            'Nenhuma NF-e de entrada em cache. '
            'Vá até a aba 🏛️ **Receita Federal** e clique em 🔄 **Buscar** para baixar os XMLs.',
            icon='📥',
        )
    else:
        _fmt_brl_e = lambda v: f'R$ {float(v):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
        with st.form('form_nf_entrada_sefaz'):
            ef1, ef2, ef3 = st.columns([1.3, 1.3, 3])
            with ef1:
                e_data_ini = st.date_input('De', value=st.session_state.data_ini,
                                           format='DD/MM/YYYY', key='e_ini')
            with ef2:
                e_data_fim = st.date_input('Até', value=st.session_state.data_fim,
                                           format='DD/MM/YYYY', key='e_fim')
            with ef3:
                e_busca = st.text_input('Buscar (fornecedor, CNPJ, número…)',
                                        key='e_busca', placeholder='Filtrar...')
            st.form_submit_button('🔍 Consultar', type='primary')

        df_e = pd.DataFrame(nfs_ent_xml)
        if 'EMISSAO' in df_e.columns:
            _d0_s = str(e_data_ini)
            _d1_s = str(e_data_fim)
            df_e = df_e[df_e['EMISSAO'].apply(
                lambda v: not v or _d0_s <= str(v)[:10] <= _d1_s
            )]
        if e_busca.strip():
            termo = e_busca.strip().lower()
            df_e = df_e[df_e.apply(lambda r: any(termo in str(v).lower() for v in r), axis=1)]

        # Enriquecer CFOP dos XMLs em cache (dados parseados antes do patch)
        _needs_cfop = ('CFOP' not in df_e.columns or
                       df_e['CFOP'].isna().all() or
                       (df_e['CFOP'].astype(str).isin(['', 'None', 'nan'])).all())
        if _needs_cfop and not df_e.empty:
            _cnpj_cfop = _limpar_cnpj(st.session_state.cert_cfg.get('cnpj', ''))
            if _cnpj_cfop:
                _cache_docs = carregar_cache(_cnpj_cfop)
                _cfop_map = {}
                for _doc in _cache_docs:
                    _schema = _doc.get('schema', '')
                    if 'procNFe' not in _schema and 'nfeProc' not in _schema:
                        continue
                    _xstr = _doc.get('xml', '')
                    try:
                        _xroot = ET.fromstring(_xstr)
                        _ns = NS_NFE
                        _inf = _xroot.find(f'.//{_ns}infNFe')
                        if _inf is None:
                            _inf = _xroot.find('.//infNFe')
                            _ns = ''
                        if _inf is None:
                            continue
                        _ch = (_inf.get('Id', '') or '').replace('NFe', '')
                        _cfops_det = []
                        for _det in _inf.findall(f'{_ns}det'):
                            _pr = _det.find(f'{_ns}prod')
                            if _pr is not None:
                                _cv = _pr.findtext(f'{_ns}CFOP') or _pr.findtext('CFOP') or ''
                                if _cv:
                                    _cfops_det.append(_cv.strip())
                        if _cfops_det:
                            _cfop_map[_ch] = max(set(_cfops_det), key=_cfops_det.count)
                    except Exception:
                        pass
                if _cfop_map and 'CHAVE' in df_e.columns:
                    df_e['CFOP'] = df_e['CHAVE'].map(_cfop_map).fillna('')

        # Classificar por CFOP
        if 'CFOP' in df_e.columns:
            df_e['TIPO'] = df_e['CFOP'].apply(_classificar_cfop)
        else:
            df_e['CFOP'] = ''
            df_e['TIPO'] = ''

        # Filtro por tipo (CFOP)
        _tipos_disp = sorted(df_e['TIPO'].unique().tolist())
        _tipos_disp = [t for t in _tipos_disp if t]
        if _tipos_disp:
            _tipo_sel = st.multiselect(
                'Filtrar por tipo (CFOP)',
                options=['Todos'] + _tipos_disp,
                default=['Todos'],
                key='filtro_tipo_cfop_ent',
            )
            if 'Todos' not in _tipo_sel and _tipo_sel:
                df_e = df_e[df_e['TIPO'].isin(_tipo_sel)]

        _cols_e = ['NUMERO', 'SERIE', 'EMISSAO', 'FORNECEDOR', 'CNPJ_EMIT', 'CFOP', 'TIPO',
                   'TOTAL_NF', 'TOTAL_PROD', 'TOTAL_IPI', 'TOTAL_ICMS', 'TOTAL_ST']
        _cols_e = [c for c in _cols_e if c in df_e.columns]
        _rename_e = {
            'NUMERO': 'Número', 'SERIE': 'Série', 'EMISSAO': 'Emissão',
            'FORNECEDOR': 'Fornecedor', 'CNPJ_EMIT': 'CNPJ Emit.',
            'CFOP': 'CFOP', 'TIPO': 'Tipo',
            'TOTAL_NF': 'Total NF', 'TOTAL_PROD': 'Produtos', 'TOTAL_IPI': 'IPI',
            'TOTAL_ICMS': 'ICMS', 'TOTAL_ST': 'ST',
        }
        _fmt_e = {c: _fmt_brl_e for c in
                  ['TOTAL_NF', 'TOTAL_PROD', 'TOTAL_IPI', 'TOTAL_ICMS', 'TOTAL_ST']
                  if c in df_e.columns}

        em1, em2, em3, em4 = st.columns(4)
        em1.metric('Total NF-e entrada', len(df_e))
        em2.metric('Total geral',
                   _fmt_brl_e(sum(df_e['TOTAL_NF'])) if 'TOTAL_NF' in df_e.columns and len(df_e) > 0 else 'R$ 0,00')
        _n_compra = len(df_e[df_e['TIPO'].str.contains('Compra', na=False)]) if 'TIPO' in df_e.columns else 0
        _n_outros = len(df_e) - _n_compra
        em3.metric('🛒 Compras', _n_compra)
        em4.metric('📋 Outros', _n_outros)

        st.dataframe(df_e[_cols_e].rename(columns=_rename_e).style.format(_fmt_e),
                     use_container_width=True, height=480)

        _col_xls, _col_xml, _col_pdf = st.columns(3)
        with _col_xls:
            st.download_button(
                label='📊 Exportar Excel (NF Entrada)',
                data=gerar_excel(nfs_ent_xml),
                file_name=f'nf_entrada_{e_data_ini.strftime("%d%m%Y")}_{e_data_fim.strftime("%d%m%Y")}.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                type='primary', key='btn_excel_entrada', use_container_width=True,
            )
        with _col_xml:
            if st.button('📦 Baixar XMLs (.zip)', key='btn_xml_entrada',
                         use_container_width=True):
                _cnpj_e = st.session_state.cert_cfg.get('cnpj', '').replace('.', '').replace('/', '').replace('-', '')
                _nfs_dict = df_e.to_dict('records') if not df_e.empty else []
                _chave_map = {n.get('CHAVE', ''): n for n in nfs_ent_xml if n.get('CHAVE')}
                _nfs_enr = [_chave_map.get(_r.get('CHAVE', ''), _r) for _r in _nfs_dict]
                with st.spinner('Montando ZIP de XMLs...'):
                    _zip_bytes, _cnt = _build_xml_zip(_nfs_enr, _cnpj_e)
                if _cnt == 0:
                    st.warning('Nenhum XML encontrado no cache para as NF-e filtradas.')
                else:
                    st.session_state['_xml_zip_bytes'] = _zip_bytes
                    st.session_state['_xml_zip_name'] = (
                        f'xmls_entrada_{e_data_ini.strftime("%d%m%Y")}_'
                        f'{e_data_fim.strftime("%d%m%Y")}.zip'
                    )
                    _total_res = sum(1 for n in _nfs_enr if n.get('RESUMO'))
                    if _total_res > 0:
                        st.info(
                            f'{_cnt} XML(s) incluídos no ZIP. '
                            f'**{_total_res} são resumos** (resNFe) — XML completo ainda não '
                            f'disponível na SEFAZ. Clique em **Buscar** periodicamente para '
                            f'baixar o XML completo quando a SEFAZ disponibilizar.',
                            icon='ℹ️',
                        )
            if st.session_state.get('_xml_zip_bytes'):
                st.download_button(
                    label='⬇️ Clique para baixar XMLs',
                    data=st.session_state['_xml_zip_bytes'],
                    file_name=st.session_state.get('_xml_zip_name', 'xmls_entrada.zip'),
                    mime='application/zip',
                    key='btn_xml_dl_entrada', use_container_width=True,
                )
        with _col_pdf:
            if st.button('📄 Gerar DANFEs (.zip)', key='btn_danfe_entrada',
                         use_container_width=True):
                _cnpj_e2 = st.session_state.cert_cfg.get('cnpj', '').replace('.', '').replace('/', '').replace('-', '')
                _chave_map2 = {n.get('CHAVE', ''): n for n in nfs_ent_xml if n.get('CHAVE')}
                _nfs_enr2 = [_chave_map2.get(_r.get('CHAVE', ''), _r)
                              for _r in (df_e.to_dict('records') if not df_e.empty else [])]
                with st.spinner('Gerando DANFEs em PDF...'):
                    _pdf_zip, _ger, _sem = _build_danfe_zip(_nfs_enr2, _cnpj_e2)
                if _ger == 0 and _sem > 0:
                    st.warning(
                        f'Nenhum DANFE gerado — todas as {_sem} NF-e do filtro são resumos SEFAZ '
                        f'(resNFe). O DANFE só pode ser gerado quando o XML completo estiver '
                        f'disponível. Clique em **Buscar** para tentar baixar os XMLs completos.',
                        icon='⚠️',
                    )
                elif _ger == 0:
                    st.warning('Nenhum DANFE gerado. Verifique os XMLs no cache.')
                else:
                    st.session_state['_pdf_zip_bytes'] = _pdf_zip
                    st.session_state['_pdf_zip_name'] = (
                        f'danfes_entrada_{e_data_ini.strftime("%d%m%Y")}_'
                        f'{e_data_fim.strftime("%d%m%Y")}.zip'
                    )
                    if _sem > 0:
                        st.info(
                            f'{_ger} DANFE(s) gerados. {_sem} NF-e são resumos SEFAZ '
                            f'(sem XML completo) e foram ignoradas.',
                            icon='ℹ️',
                        )
            if st.session_state.get('_pdf_zip_bytes'):
                st.download_button(
                    label='⬇️ Clique para baixar DANFEs',
                    data=st.session_state['_pdf_zip_bytes'],
                    file_name=st.session_state.get('_pdf_zip_name', 'danfes_entrada.zip'),
                    mime='application/zip',
                    key='btn_pdf_dl_entrada', use_container_width=True,
                )

# ══════════════════════════════════════════════════════════════════════════════
# TAB: CUPONS NFC-e / SAT
# ══════════════════════════════════════════════════════════════════════════════
with tab_cupons:
    st.subheader('Cupons Fiscais — NFC-e (modelo 65) e SAT CFe')
    st.caption('Os cupons são lidos diretamente do banco ERP (tabela VENDAS).')

    cup_col1, cup_col2, cup_col3, cup_col4, cup_col5 = st.columns([1.3, 1.3, 2, 1, 1])
    with cup_col1:
        cup_data_ini = st.date_input('De', value=st.session_state.data_ini,
                                     format='DD/MM/YYYY', key='cup_ini')
    with cup_col2:
        cup_data_fim = st.date_input('Até', value=st.session_state.data_fim,
                                     format='DD/MM/YYYY', key='cup_fim')
    with cup_col3:
        c_busca = st.text_input('Buscar (número, CNPJ…)', key='c_busca', placeholder='Filtrar...')
    with cup_col4:
        btn_ler_cupons_db = st.button('🔄 Carregar do ERP', type='primary', key='btn_ler_cupons_db')
    with cup_col5:
        pasta_nfce_cfg = st.session_state.cert_cfg.get('pasta_nfce', '')
        btn_ler_cupons_pasta = st.button('📂 Carregar da pasta', key='btn_ler_cupons_pasta',
                                         disabled=not bool(pasta_nfce_cfg))

    if btn_ler_cupons_db:
        db_cfg = st.session_state.db_cfg
        if not db_cfg:
            st.error('Configure a conexão com o ERP na aba ⚙️ Configurações.')
        else:
            with st.spinner('Consultando banco ERP...'):
                try:
                    con_cup = conectar_erp_loja(_loja_ativa, db_cfg)
                    _db_local_cup = bool(_loja_ativa and _loja_ativa.get('db_host'))
                    _cnpj_loja_cup = _limpar_cnpj(st.session_state.cert_cfg.get('cnpj', ''))
                    cupons_lidos, erros_cupons = ler_cupons_banco(con_cup, cup_data_ini, cup_data_fim,
                                                                   cnpj_loja=None if _db_local_cup else (_cnpj_loja_cup or None))
                    con_cup.close()
                    st.session_state.cupons_mem = cupons_lidos
                    if erros_cupons:
                        with st.expander(f'⚠️ {len(erros_cupons)} aviso(s)'):
                            for e in erros_cupons:
                                st.text(e)
                    st.success(f'✔ {len(cupons_lidos)} cupons carregados do ERP.')
                except Exception as ex:
                    st.error(f'Erro ao conectar ao ERP: {ex}')

    if btn_ler_cupons_pasta:
        if not pasta_nfce_cfg or not os.path.isdir(pasta_nfce_cfg):
            st.error('Pasta não encontrada. Configure em Configurações.')
        else:
            with st.spinner('Lendo XMLs da pasta...'):
                cupons_lidos, erros_cupons = ler_cupons_pasta(pasta_nfce_cfg)
                st.session_state.cupons_mem = cupons_lidos
                if erros_cupons:
                    with st.expander(f'⚠️ {len(erros_cupons)} arquivo(s) com erro'):
                        for e in erros_cupons:
                            st.text(e)
                if cupons_lidos:
                    st.success(f'✔ {len(cupons_lidos)} cupons carregados da pasta.')

    cupons = st.session_state.cupons_mem
    if cupons:
        df_c = pd.DataFrame(cupons)
        if 'EMISSAO' in df_c.columns:
            df_c['_dt'] = pd.to_datetime(df_c['EMISSAO'], errors='coerce')
            df_c = df_c[(df_c['_dt'].isna()) |
                        ((df_c['_dt'] >= pd.Timestamp(cup_data_ini)) &
                         (df_c['_dt'] <= pd.Timestamp(cup_data_fim)))]
            df_c = df_c.drop(columns=['_dt'])
        if c_busca.strip():
            termo = c_busca.strip().lower()
            df_c = df_c[df_c.apply(lambda r: any(termo in str(v).lower() for v in r), axis=1)]

        _fmt_brl_c = lambda v: f'R$ {float(v):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')

        # Marca coluna Situação
        if 'CANCELADO' in df_c.columns:
            df_c['_situacao'] = df_c['CANCELADO'].apply(lambda x: '❌ Cancelado' if x else '✅ Normal')
        else:
            df_c['_situacao'] = '✅ Normal'

        n_normais   = len(df_c[df_c['_situacao'] == '✅ Normal'])
        n_cancelados = len(df_c[df_c['_situacao'] == '❌ Cancelado'])
        nfce_cnt = len(df_c[df_c['TIPO'] == 'NFC-e']) if 'TIPO' in df_c.columns else 0
        sat_cnt  = len(df_c[df_c['TIPO'] == 'SAT'])   if 'TIPO' in df_c.columns else 0

        total_todos   = sum(df_c['TOTAL_NF']) if 'TOTAL_NF' in df_c.columns and len(df_c) > 0 else 0
        total_canc    = sum(df_c.loc[df_c['_situacao'] == '❌ Cancelado', 'TOTAL_NF']) if n_cancelados > 0 else 0
        total_liquido = total_todos - total_canc

        cm1, cm2, cm3, cm4 = st.columns(4)
        cm1.metric('Total cupons', len(df_c))
        cm2.metric('NFC-e', nfce_cnt)
        cm3.metric('SAT', sat_cnt)
        cm4.metric('❌ Cancelados', n_cancelados)

        tv1, tv2, tv3 = st.columns(3)
        tv1.metric('Total geral (bruto)', _fmt_brl_c(total_todos))
        tv2.metric('Total cancelados', _fmt_brl_c(total_canc))
        tv3.metric('Total líquido', _fmt_brl_c(total_liquido))

        _cols_c = [c for c in ['_situacao', 'TIPO', 'NUMERO', 'EMISSAO', 'TOTAL_NF'] if c in df_c.columns]
        _rename_c = {'_situacao': 'Situação', 'TIPO': 'Tipo',
                     'NUMERO': 'Número', 'EMISSAO': 'Emissão', 'TOTAL_NF': 'Total'}
        _fmt_c = {'TOTAL_NF': _fmt_brl_c}

        def _colorir_cupons(row):
            if '❌' in str(row.get('Situação', '')):
                return ['background-color: #FCE4D6'] * len(row)
            return [''] * len(row)

        st.dataframe(
            df_c[_cols_c].rename(columns=_rename_c).style.apply(_colorir_cupons, axis=1).format(_fmt_c),
            use_container_width=True, height=480,
        )

        _cup_dl1, _cup_dl2 = st.columns(2)
        with _cup_dl1:
            _rows_cup_exp = [{k: v for k, v in r.items() if k != 'XML'} for r in cupons]
            st.download_button(
                label='📊 Exportar Excel',
                data=gerar_excel(_rows_cup_exp),
                file_name=f'cupons_{cup_data_ini.strftime("%d%m%Y")}_{cup_data_fim.strftime("%d%m%Y")}.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                type='primary', key='btn_excel_cupons', use_container_width=True,
            )
        with _cup_dl2:
            _xmls_cup = [r.get('XML', '') for r in cupons if r.get('XML')]
            if _xmls_cup:
                import zipfile as _zfc, io as _ioc
                _zbuf_c = _ioc.BytesIO()
                with _zfc.ZipFile(_zbuf_c, 'w', _zfc.ZIP_DEFLATED) as _zzc:
                    for _ic, _cr in enumerate(cupons):
                        if _cr.get('XML'):
                            _ch_c = _cr.get('CHAVE', '') or f'cupom_{_ic+1:04d}'
                            _zzc.writestr(f'{_ch_c}.xml', _cr['XML'])
                st.download_button(
                    f'📦 Baixar XMLs ({len(_xmls_cup)})',
                    data=_zbuf_c.getvalue(),
                    file_name=f'xmls_cupons_{cup_data_ini.strftime("%d%m%Y")}_{cup_data_fim.strftime("%d%m%Y")}.zip',
                    mime='application/zip',
                    key='btn_zip_cupons', use_container_width=True,
                )
            else:
                st.button('📦 XMLs no banco local da loja', disabled=True, key='btn_zip_cupons_dis',
                          use_container_width=True,
                          help='Os XMLs dos cupons ficam no banco local de cada loja. Configure o IP da loja em Configurações > Banco de dados por loja.')
    else:
        st.info('Clique em "Carregar cupons da pasta" para importar os XMLs.', icon='ℹ️')

# ══════════════════════════════════════════════════════════════════════════════
# TAB: CONTABILIDADE
# ══════════════════════════════════════════════════════════════════════════════
with tab_contab:
    st.subheader('📦 Exportar pacote para Contabilidade')
    st.write(
        'Gera um arquivo Excel com múltiplas abas (NF Entrada, NF Saída, Cupons, Conferências e Resumo) '
        'e um ZIP com todos os XMLs organizados em pastas.'
    )

    cont_col1, cont_col2 = st.columns(2)
    with cont_col1:
        cont_ini = st.date_input('Período — De', value=st.session_state.data_ini,
                                 format='DD/MM/YYYY', key='cont_ini')
    with cont_col2:
        cont_fim = st.date_input('Até', value=st.session_state.data_fim,
                                 format='DD/MM/YYYY', key='cont_fim')

    # Filtra os dados pelo período
    def _filtrar_periodo(lista, campo='EMISSAO'):
        out = []
        for item in lista:
            val = str(item.get(campo, '') or '')[:10]
            if val and str(cont_ini) <= val <= str(cont_fim):
                out.append(item)
            elif not val:
                out.append(item)
        return out

    ent_cont = _filtrar_periodo(st.session_state.nfs_xml_mem)
    sai_cont = _filtrar_periodo(st.session_state.nfs_xml_saida_mem)
    cup_cont = _filtrar_periodo(st.session_state.cupons_mem)
    res_ent  = st.session_state.resultados
    res_sai  = st.session_state.resultados_saida

    _fmt_brl_cont = lambda v: f'R$ {float(v):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')

    cm1, cm2, cm3, cm4 = st.columns(4)
    cm1.metric('NF Entrada', len(ent_cont))
    cm2.metric('NF Saída', len(sai_cont))
    cm3.metric('Cupons', len(cup_cont))
    cm4.metric('Conf. Saída', len(res_sai))

    st.info(
        '💡 Dica: para obter as conferências no pacote, realize a conferência nas abas '
        '**📋 Conf. NF Entrada** (entrada) e **📤 NF de Saída** (saída) antes de exportar.',
        icon='💡',
    )

    st.markdown('---')
    cb1, cb2 = st.columns(2)

    with cb1:
        st.write('**📊 Excel para o contador** (todas as abas em um arquivo)')
        if st.button('📊 Gerar Excel Contabilidade', key='btn_gerar_excel_cont', type='primary'):
            with st.spinner('Gerando Excel...'):
                try:
                    xls_cont = gerar_excel_contabilidade(
                        ent_cont, sai_cont, cup_cont, res_ent, res_sai, cont_ini, cont_fim
                    )
                    st.session_state['_xls_cont'] = xls_cont
                    st.success('Excel gerado!')
                except Exception as ex:
                    st.error(f'Erro: {ex}')
        if '_xls_cont' in st.session_state and st.session_state['_xls_cont']:
            st.download_button(
                '⬇️ Baixar Excel',
                data=st.session_state['_xls_cont'],
                file_name=f'contabilidade_{cont_ini.strftime("%Y%m")}_{cont_fim.strftime("%Y%m")}.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                key='btn_dl_xls_cont',
            )

    with cb2:
        st.write('**🗜️ ZIP com XMLs** (Entrada/, Saida/, Cupons/ organizados)')
        if st.button('🗜️ Gerar ZIP XMLs', key='btn_gerar_zip', type='primary'):
            with st.spinner('Empacotando XMLs...'):
                try:
                    cnpj_zip = _limpar_cnpj(st.session_state.cert_cfg.get('cnpj', ''))
                    docs_ent_raw = filtrar_nfes(carregar_cache(cnpj_zip), cnpj_zip) if cnpj_zip else []
                    docs_sai_raw = filtrar_nfes_saida(carregar_cache(cnpj_zip), cnpj_zip) if cnpj_zip else []
                    pasta_cup = st.session_state.cert_cfg.get('pasta_nfce', '')
                    zip_data = gerar_zip_xmls(docs_ent_raw, docs_sai_raw, pasta_cup)
                    st.session_state['_zip_cont'] = zip_data
                    st.success('ZIP gerado!')
                except Exception as ex:
                    st.error(f'Erro: {ex}')
        if '_zip_cont' in st.session_state and st.session_state['_zip_cont']:
            st.download_button(
                '⬇️ Baixar ZIP',
                data=st.session_state['_zip_cont'],
                file_name=f'xmls_fiscais_{cont_ini.strftime("%Y%m")}_{cont_fim.strftime("%Y%m")}.zip',
                mime='application/zip',
                key='btn_dl_zip_cont',
            )

    # ── Envio por e-mail — um ZIP por loja ────────────────────────────────────
    st.markdown('---')
    st.subheader('📧 Enviar XMLs por e-mail para a Contabilidade')
    st.write(
        'Gera um ZIP por loja (com Entrada/, Saida/, Cupons/) e envia tudo em um único e-mail.'
    )

    _ecfg_cont = load_email_config()
    _email_conf_ok = bool(
        _ecfg_cont.get('remetente') and
        _ecfg_cont.get('senha_app') and
        _ecfg_cont.get('destinatario')
    )

    if not _email_conf_ok:
        st.warning(
            '⚙️ Configure as credenciais de e-mail na aba **⚙️ Configurações** '
            '(seção "📧 E-mail para Contabilidade") antes de enviar.',
            icon='⚠️',
        )
    else:
        _mes_ref_disp = cont_ini.strftime('%B/%Y')
        _emails_adicionais = 'arthur.mataveli@controltech.com.br, blink_rh@lojasblink.com.br'
        st.info(
            f'📬 Destinatário: **{_ecfg_cont["destinatario"]}**   '
            f'| Remetente: `{_ecfg_cont["remetente"]}`\n\n'
            f'**Cópia para:** {_emails_adicionais}\n\n'
            f'Os ZIPs serão nomeados como `NOME_LOJA_{cont_ini.strftime("%Y_%m")}.zip`.',
            icon='📧',
        )

        # ── Verificação prévia de conexão com as lojas ────────────────
        _lojas_pre = load_lojas_config()
        _lojas_com_db = [l for l in _lojas_pre if l.get('db_host')]
        if st.button('🔍 Verificar conexão das lojas antes de enviar', key='btn_pre_check_lojas'):
            _ok_list = []
            _fail_list = []
            _prog_check = st.progress(0, text='Testando conexões...')
            for _ic, _lc in enumerate(_lojas_com_db):
                _prog_check.progress((_ic) / len(_lojas_com_db), text=f'Testando: {_lc["nome"]}...')
                try:
                    _cfg_test = {
                        'host': _lc['db_host'], 'port': int(_lc.get('db_port', 3050)),
                        'database': _lc['db_database'], 'user': _lc.get('db_user', 'SYSDBA'),
                        'password': _lc.get('db_password', ''),
                    }
                    _conn_test = conectar_erp(_cfg_test)
                    _conn_test.close()
                    _ok_list.append(_lc['nome'])
                except Exception as _ec:
                    _fail_list.append((_lc['nome'], str(_ec)))
            _prog_check.empty()
            if not _fail_list:
                st.success(f'✅ Todas as {len(_ok_list)} lojas com banco remoto estão online!')
                st.session_state['_lojas_check_ok'] = True
                st.session_state['_lojas_check_falhas'] = []
            else:
                st.session_state['_lojas_check_ok'] = False
                st.session_state['_lojas_check_falhas'] = _fail_list
                st.error(f'❌ {len(_fail_list)} loja(s) sem conexão:')
                for _fn, _fe in _fail_list:
                    st.caption(f'⚠️ **{_fn}** — {_fe}')
                if _ok_list:
                    st.info(f'✅ Lojas online: {", ".join(_ok_list)}')

        # Mostrar aviso persistente se houve falhas na última verificação
        _check_falhas = st.session_state.get('_lojas_check_falhas', [])
        if _check_falhas:
            _nomes_falha = [f[0] for f in _check_falhas]
            st.warning(
                f'⚠️ **Atenção:** as lojas **{", ".join(_nomes_falha)}** não estão acessíveis. '
                f'Se enviar agora, os XMLs dessas lojas **não serão incluídos** no e-mail ao contador. '
                f'Verifique a internet/Tailscale dessas lojas antes de prosseguir.',
                icon='⚠️',
            )
            _confirmar_envio = st.checkbox(
                'Estou ciente e desejo enviar mesmo sem as lojas acima',
                key='chk_confirmar_envio_parcial', value=False)
        else:
            _confirmar_envio = True

        _pode_enviar = (_confirmar_envio and
                        (st.session_state.get('_lojas_check_ok', False) or
                         st.session_state.get('_lojas_check_falhas') is not None))

        if st.button('📧 Gerar e Enviar ZIPs por Loja', key='btn_env_email_cont', type='primary',
                     disabled=not st.session_state.get('_lojas_check_ok', False) and not _confirmar_envio):
            _lojas_email = load_lojas_config()
            _zips_prontos = []
            _avisos_email = []
            _prog_email = st.progress(0, text='Preparando...')

            for _i_e, _loja_e in enumerate(_lojas_email):
                _prog_email.progress(
                    (_i_e) / len(_lojas_email),
                    text=f'Gerando ZIP: {_loja_e["nome"]}...',
                )
                try:
                    _zip_b = gerar_zip_loja_bytes(
                        _loja_e, st.session_state.db_cfg, cont_ini, cont_fim
                    )
                    if _zip_b:
                        _slug = _slug_loja(_loja_e['nome'])
                        _nome_zip = f'{_slug}_{cont_ini.strftime("%Y_%m")}.zip'
                        _zips_prontos.append((_nome_zip, _zip_b))
                    else:
                        _avisos_email.append(f'Loja {_loja_e["nome"]}: sem dados no período (ZIP não gerado).')
                except Exception as _ez:
                    _avisos_email.append(f'Loja {_loja_e["nome"]}: erro ao gerar ZIP — {_ez}')

            _prog_email.progress(1.0, text='Enviando e-mail...')

            if _zips_prontos:
                try:
                    enviar_email_contabilidade(
                        _zips_prontos,
                        _ecfg_cont['remetente'],
                        _ecfg_cont.get('senha_app', ''),
                        _ecfg_cont['destinatario'],
                        _mes_ref_disp,
                        brevo_api_key=_ecfg_cont.get('brevo_api_key', ''),
                    )
                    _prog_email.empty()
                    _metodo_env = 'Brevo' if _ecfg_cont.get('brevo_api_key') else 'Gmail SMTP'
                    st.success(
                        f'✔ E-mail enviado via {_metodo_env} para **{_ecfg_cont["destinatario"]}** '
                        f'com {len(_zips_prontos)} ZIP(s) em anexo!'
                    )
                    # Lista os ZIPs enviados
                    for _nz, _dz in _zips_prontos:
                        st.caption(f'📎 {_nz}  ({len(_dz)/1024:.0f} KB)')
                except Exception as _ee:
                    _prog_email.empty()
                    st.error(f'Erro ao enviar e-mail: {_ee}')
            else:
                _prog_email.empty()
                st.warning('Nenhum ZIP gerado — sem dados NF-e para o período em nenhuma loja.')

            if _avisos_email:
                with st.expander(f'⚠️ Avisos ({len(_avisos_email)})'):
                    for _av in _avisos_email:
                        st.caption(_av)

    # ── Status dos E-mails Enviados (Brevo) ──────────────────────────────────
    st.markdown('---')
    st.subheader('📧 Status dos E-mails Enviados')
    _ecfg_status = load_email_config()
    if not _ecfg_status.get('brevo_api_key'):
        st.info('Configure a **Brevo API Key** em ⚙️ Configurações para ver o status dos e-mails enviados.')
    else:
        _dias_filtro = st.selectbox('Período', [7, 15, 30, 60, 90],
                                    index=2, format_func=lambda d: f'Últimos {d} dias',
                                    key='brevo_dias_filtro')
        if st.button('🔄 Consultar status', key='btn_consultar_brevo'):
            with st.spinner('Consultando Brevo...'):
                try:
                    _status_list = consultar_status_emails_brevo(
                        _ecfg_status['brevo_api_key'], dias=_dias_filtro)
                    if _status_list:
                        import pandas as _pd_brevo
                        _df_status = _pd_brevo.DataFrame(_status_list)
                        # Contadores
                        _total = len(_df_status)
                        _abertos = len(_df_status[_df_status['Status'].str.contains('Aberto')])
                        _entregues = len(_df_status[_df_status['Status'].str.contains('Entregue')])
                        _c1, _c2, _c3, _c4 = st.columns(4)
                        _c1.metric('Total enviados', _total)
                        _c2.metric('📨 Entregues', _entregues)
                        _c3.metric('✅ Abertos', _abertos)
                        _c4.metric('📊 Taxa abertura',
                                   f'{(_abertos/_total*100):.0f}%' if _total else '0%')
                        st.dataframe(_df_status, use_container_width=True, hide_index=True)
                    else:
                        st.info('Nenhum e-mail encontrado no período selecionado.')
                except Exception as _eb:
                    st.error(f'Erro ao consultar Brevo: {_eb}')
